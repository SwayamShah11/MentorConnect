import traceback
import uuid
import base64
from django.core.mail import send_mail
from django.shortcuts import render, redirect, get_object_or_404
import io
from django.utils.decorators import method_decorator
import os
import re
from urllib.parse import urlencode
from ..models import WEEK_CHOICES, YEAR_CHOICES, SEM_CHOICES
from django.views.decorators.http import require_POST
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from html import escape
from django.contrib import messages
from ..forms import MentorRegisterForm, MentorProfileForm, MoodleIdForm, ChatReplyForm, MentorInteractionForm, WeeklyAgendaForm
from django.views.generic import (View, TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView)
from django.utils.timezone import now
from django.db.models.functions import TruncMonth
from django.db.models import Count, Q, Avg, Sum
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.views.generic import TemplateView
from ..models import (Profile, Msg, Conversation, Reply, Meeting, Mentor, Mentee, MentorMentee, Query, InternshipPBL,
                      PaperPublication, SemesterResult, SportsCulturalEvent, CertificationCourse, OtherEvent, Project,
                      MentorMenteeInteraction, Notification, ReminderLog, WeeklyAgenda)
from ..utils import get_document_progress, mentor_required, mentor_or_staff_required
from django.views.decorators.csrf import csrf_exempt
from mentee.ai_utils import generate_ai_summary
from datetime import datetime, timedelta
from django.utils import timezone
from django.urls import reverse_lazy
from ..forms import ReplyForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.mixins import UserPassesTestMixin
from reportlab.lib.styles import ParagraphStyle
from django.contrib.auth import get_user_model

User = get_user_model()
from django.contrib.messages.views import SuccessMessageMixin
from django.core.paginator import Paginator
import json
import zipfile
from io import BytesIO
from datetime import datetime
from collections import Counter, defaultdict

from django.http import HttpResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.conf import settings

# Excel libs
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

# ReportLab libs
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (SimpleDocTemplate, Table as RLTable, TableStyle as RLTableStyle, Paragraph, Spacer, Image as RLImage, PageBreak)
from reportlab.lib import colors
from reportlab.lib.units import cm
from itertools import chain
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot

SEMESTER = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']       #used in Mentor-Mentee interaction function


def home(request):
    """Landing page """
    return render(request, 'home.html')

#-----------------Mentor dashboard logic starts-------------------
@method_decorator(login_required, name='dispatch')
class AccountView(LoginRequiredMixin, UserPassesTestMixin, View):
    """For Mentor Account"""

    def test_func(self):
        # Only allow if the logged-in user is a mentor
        return hasattr(self.request.user, "mentor")

    def handle_no_permission(self):
        # If user is logged in but not mentor â†’ redirect to home
        if self.request.user.is_authenticated:
            return redirect("home")
        # If not logged in â†’ go to login
        return super().handle_no_permission()

    def get(self, request):
        # --- MESSAGE REMINDERS ---
        unapproved_messages = Msg.objects.filter(
            receipient=request.user,
            is_approved=False
        ).count()

        # --- QUERY REMINDERS ---
        mentor = get_object_or_404(Mentor, user=request.user)
        pending_queries = Query.objects.filter(
            mentor=mentor,
            status__in=["pending", "open"]
        ).count()

        # --- MEETING REMINDERS ---
        pending_meetings = [
            m for m in Meeting.objects.filter(mentor=mentor, status='scheduled')
            if m.meeting_end_datetime > timezone.localtime(timezone.now())
        ]
        pending_meetings_count = len(pending_meetings)

        form = MoodleIdForm()
        mentor = get_object_or_404(Mentor, user=request.user)

        mentee_mappings = MentorMentee.objects.filter(mentor=mentor).select_related("mentee__user")

        mentees = []
        no_document_mentees = []

        for mapping in mentee_mappings:
            user = mapping.mentee.user
            profile = Profile.objects.filter(user=user).first()

            completed_count, total_required, has_pending = get_document_progress(user)
            progress_percent = int((completed_count / total_required) * 100)

            mentee_data = {
                "id": mapping.mentee.pk,
                "moodle_id": profile.moodle_id if profile else "",
                "name": profile.student_name if profile else user.username,
                "semester": profile.semester if profile else "",
                "contact": profile.contact_number if profile else "",
                "progress": f"{completed_count}/{total_required}",
                "progress_percent": progress_percent,
                "has_any_document": completed_count > 0,
                "has_pending": has_pending,
            }

            mentees.append(mentee_data)

            if completed_count < 7:
                no_document_mentees.append(mentee_data)
        pending_reminder_count = len(no_document_mentees)
        return render(request, "mentor/account1.html", {
            "form": form,
            "mentees": mentees,
            "no_document_mentees": no_document_mentees,
            "pending_reminder_count": pending_reminder_count,
            "unapproved_messages": unapproved_messages,
            "pending_queries": pending_queries,
            "pending_meetings": pending_meetings_count,
        })

    def post(self, request):
        form = MoodleIdForm(request.POST)
        mentor = get_object_or_404(Mentor, user=request.user)

        if form.is_valid():
            moodle_id = form.cleaned_data["moodle_id"]
            profile = Profile.objects.filter(moodle_id=moodle_id).first()

            if not profile:
                messages.error(request, "No student found with that Moodle ID.")
                return redirect("account1")

            mentee = Mentee.objects.filter(user=profile.user).first()
            if not mentee:
                mentee = Mentee.objects.create(user=profile.user)

            # Prevent duplicates
            # âœ… PREVENT MULTIPLE MENTORS ASSIGNING SAME MENTEE
            existing_mapping = MentorMentee.objects.filter(mentee=mentee).first()

            if existing_mapping:
                messages.error(request,f"{moodle_id}-{profile.student_name} is already assigned to Mentor: {existing_mapping.mentor.name}")
                return redirect("account1")

            MentorMentee.objects.create(mentor=mentor, mentee=mentee)
            messages.success(request, f"{moodle_id}-{profile.student_name} added as your mentee!")

        return redirect("account1")


@login_required
def remind_mentee(request, mentee_id):
    mentor = get_object_or_404(Mentor, user=request.user)
    mentee = get_object_or_404(Mentee, pk=mentee_id)
    user = mentee.user

    completed_count, total_required, has_pending = get_document_progress(user)

    if not has_pending:
        messages.info(request, f"{user.username} - {user.profile.student_name} has already completed all required documents.")
        return redirect("account1")

    # âœ… 1. Save Notification
    Notification.objects.create(
        user=user,
        message=f"âš ï¸ Your mentor {mentor.name} has reminded you to upload your pending documents and complete the Profile.\n"
    )

    # âœ… 2. Send Email
    if user.email:
        send_mail(
            subject="Reminder to Upload Your Documents and Complete the Profile",
            message=(
                f"Dear {user.profile.student_name}({user.username}),\n\n"
                f"Your mentor {mentor.name} has reminded you to upload your pending documents and complete your profile.\n"
                f"You need to complete the pending uploads before the next mentoring session.\n"
                f"Please log in to MentorConnect and upload them as soon as possible.\n\n"
                f"Regards,\nMentorConnect Team\n\n\n"
                f"*This is a system generated Email. Please do not reply to this Email.*\n"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    # 3ï¸âƒ£ Log this reminder
    ReminderLog.objects.create(
        mentor=mentor,
        mentee=mentee,
        is_auto=False,
    )
    messages.success(request, f"Reminder sent to {user.username} - {user.profile.student_name}!")
    return redirect("account1")


@login_required
def remind_all_mentees(request):
    mentor = get_object_or_404(Mentor, user=request.user)

    mentee_mappings = MentorMentee.objects.filter(mentor=mentor).select_related(
        "mentee__user", "mentee__user__profile"
    )

    reminded_count = 0

    for mapping in mentee_mappings:
        mentee = mapping.mentee
        user = mentee.user

        completed_count, total_required, has_pending = get_document_progress(user)

        # Skip mentees with no pending documents
        if not has_pending:
            continue

        # âœ… 1. Save Notification
        Notification.objects.create(
            user=user,
            message=(
                f"âš ï¸ Your mentor {mentor.name} has reminded you to upload your "
                f"pending documents and complete the Profile.\n"
            )
        )

        # âœ… 2. Send Email
        if user.email:
            send_mail(
                subject="Reminder to Upload Your Documents and Complete the Profile",
                message=(
                    f"Dear {user.profile.student_name} ({user.username}),\n\n"
                    f"Your mentor {mentor.name} has reminded you to upload your pending "
                    f"documents and complete your profile.\n"
                    f"You need to complete the pending uploads before the next mentoring session.\n\n"
                    f"Please log in to MentorConnect and upload them as soon as possible.\n\n"
                    f"Regards,\nMentorConnect Team\n\n"
                    f"*This is a system generated Email. Please do not reply to this Email.*\n"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                fail_silently=True,
            )

        # âœ… 3. Log Reminder
        ReminderLog.objects.create(
            mentor=mentor,
            mentee=mentee,
            is_auto=False,
        )

        reminded_count += 1

    # âœ… Feedback message
    if reminded_count:
        messages.success(request, f"Reminder sent to {reminded_count} mentees successfully!")
    else:
        messages.info(request, "All mentees have already completed their documents.")

    return redirect("account1")
#-----------------Mentor dashboard logic ends-------------------


def register1(request):
    """Registration for mentors"""

    registered = False

    if request.method == 'POST':

        form1 = MentorRegisterForm(request.POST)

        if form1.is_valid():
            user = form1.save()
            user.is_mentor = True
            user.save()

            registered = True
            messages.success(request, f'Your account has been created! You are now able to log in')
            return redirect('login1')
    else:
        form1 = MentorRegisterForm()

    return render(request, 'mentor/register1.html', {'form1': form1})


#---------------Profile page logic starts-----------------
@login_required
def profile1(request):
    """Update Mentor Profile"""

    if not hasattr(request.user, "mentor"):
        return redirect("home")

    mentor = request.user.mentor  # one-to-one relation

    if request.method == "POST":
        form = MentorProfileForm(request.POST, request.FILES, instance=mentor, user=request.user)

        if form.is_valid():
            # Update User info
            request.user.username = form.cleaned_data["username"]
            request.user.email = form.cleaned_data["email"]
            request.user.save()

            # Update Mentor info
            form.save()

            messages.success(request, "Your profile has been updated successfully!")
            return redirect("profile1")
    else:
        form = MentorProfileForm(instance=mentor, user=request.user)

    return render(request, "mentor/profile1.html", {"form": form})
#---------------Profile page logic ends-----------------


def user_login(request):
    """Login function"""

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            if user.is_active:
                login(request, user)
                messages.success(request, f'Welcome to your Account {user}')
                return HttpResponseRedirect(reverse('account1'))
            else:
                return HttpResponse("Your account was inactive.")
        else:
            print("Someone tried to login and failed.")
            print("They used username: {} and password: {}".format(username, password))
            messages.warning(request, f'Invalid login details')
            return redirect("login1")
    else:
        return render(request, 'mentor/login1.html', {})


def custom_logout(request):
    logout(request)
    return redirect('login1')


@login_required
def remove_mentee(request, mentee_id):
    if not hasattr(request.user, "mentor"):
        return redirect("home")  # block non-mentors

    mentor = get_object_or_404(Mentor, user=request.user)
    mapping = get_object_or_404(MentorMentee, mentor=mentor, mentee_id=mentee_id)
    mapping.delete()
    return redirect("account1")


@login_required
def view_mentee(request, mentee_id):
    from mentee.models import (
        EducationalDetail, InternshipPBL,Project, SportsCulturalEvent, OtherEvent, LongTermGoal, CertificationCourse,
        PaperPublication, SelfAssessment, StudentInterest, SemesterResult, Profile, SubjectOfInterest
    )

    # get mentee profile
    mentee = get_object_or_404(User, pk=mentee_id)

    # fetch related objects
    profile = Profile.objects.filter(user=mentee).first()
    education = EducationalDetail.objects.filter(user=mentee)
    projects = Project.objects.filter(user=mentee)
    sports = SportsCulturalEvent.objects.filter(user=mentee)
    other_event = OtherEvent.objects.filter(user=mentee)
    publications = PaperPublication.objects.filter(user=mentee)
    student_interest = StudentInterest.objects.filter(student=mentee).first()
    results = SemesterResult.objects.filter(user=mentee)
    internships = InternshipPBL.objects.filter(user=mentee)
    goals = LongTermGoal.objects.filter(user=mentee)
    subjects = SubjectOfInterest.objects.filter(user=mentee)
    certifications = CertificationCourse.objects.filter(user=mentee)
    assessment = SelfAssessment.objects.filter(user=mentee)

    context = {
        "mentee": mentee,
        "profile": profile,
        "education": education,
        "projects": projects,
        "sports": sports,
        "other_event": other_event,
        "publications": publications,
        "student_interest": student_interest,
        "saved_interests": student_interest.interests if student_interest else [],
        "results": results,
        "internships": internships,
        "goal": goals,
        "subjects": subjects,
        "certifications": certifications,
        "assessment": assessment,
        "is_mentor_view": True,   # ðŸ‘ˆ flag to hide edit buttons
    }
    return render(request, "mentor/view_mentee_dashboard.html", context)


#-------------------Download mentee docs logic starts----------------------
@login_required
def mentee_documents(request):
    documents = []
    error = None
    selected_moodles = []
    selected_mentees_info = []  # [{moodle_id, name}...]

    mentor = get_object_or_404(Mentor, user=request.user)

    mentee_mappings = MentorMentee.objects.filter(mentor=mentor).select_related("mentee__user")
    mentees = []
    mentee_map = {}  # moodle_id(str) -> {"user": user, "name": name}

    for mapping in mentee_mappings:
        profile = Profile.objects.filter(user=mapping.mentee.user).first()
        if not profile or not profile.moodle_id:
            continue

        moodle_str = str(profile.moodle_id).strip()
        name = (profile.student_name or mapping.mentee.user.username)

        mentees.append({
            "user_id": mapping.mentee.user.id,
            "moodle_id": moodle_str,
            "name": name,
        })
        mentee_map[moodle_str] = {"user": mapping.mentee.user, "name": name}

    if request.method == "POST":
        selected_moodles = [m.strip() for m in request.POST.getlist("moodle_ids") if m.strip()]

        if not selected_moodles:
            error = "Please select at least one Moodle ID."
        else:
            invalid = [m for m in selected_moodles if m not in mentee_map]
            if invalid:
                error = "Selected Moodle ID is not one of your mentees."
            else:
                # Helper
                def add_doc(owner_moodle, owner_name, file, label, dtype):
                    if file:
                        documents.append({
                            "owner_moodle": owner_moodle,
                            "owner_name": owner_name,
                            "name": label,
                            "file": file,
                            "type": dtype,
                        })

                # Build docs for each selected mentee
                for moodle_id in selected_moodles:
                    mentee_user = mentee_map[moodle_id]["user"]
                    mentee_name = mentee_map[moodle_id]["name"]

                    selected_mentees_info.append({
                        "moodle_id": moodle_id,
                        "name": mentee_name,
                    })

                    for item in InternshipPBL.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.certificate,
                                item.title or "Internship / PBL Certificate", "Internship / PBL")

                    for item in SportsCulturalEvent.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.certificate,
                                item.name_of_event or "Sports / Cultural Event", "Sports / Cultural")

                    for item in OtherEvent.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.certificate,
                                item.name_of_event or "Other Event", "Other Event")

                    for item in CertificationCourse.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.certificate,
                                item.title or "Certification Course", "Course")

                    for item in PaperPublication.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.certificate,
                                item.title or "Paper Publication", "Publication")

                    for item in SemesterResult.objects.filter(user=mentee_user):
                        add_doc(moodle_id, mentee_name, item.marksheet,
                                f"Semester {item.semester} Marksheet", "Semester Result")

                # Optional: stable ordering (group by mentee then type then name)
                documents.sort(key=lambda d: (d["owner_moodle"], d["type"], (d["name"] or "")))

    return render(request, "mentor/mentee_documents.html", {
        "documents": documents,
        "mentees": mentees,
        "error": error,
        "selected_moodles": selected_moodles,
        "selected_mentees_info": selected_mentees_info,
    })


@login_required
def download_all_documents(request):
    if request.method != "POST":
        return redirect("mentee_documents")

    mentor = get_object_or_404(Mentor, user=request.user)
    mappings = MentorMentee.objects.filter(mentor=mentor).select_related("mentee__user")

    # Build {moodle_id: user}
    mentees = {}
    for mapping in mappings:
        profile = Profile.objects.filter(user=mapping.mentee.user).first()
        if profile and profile.moodle_id:
            mentees[str(profile.moodle_id).strip()] = mapping.mentee.user

    download_scope = request.POST.get("download_scope", "selected")  # "all" or "selected"
    selected_moodles = [m.strip() for m in request.POST.getlist("moodle_ids") if m.strip()]  # multi-select
    selected_types = request.POST.getlist("doc_types")     # optional multi-select

    if download_scope == "all":
        target_moodles = list(mentees.keys())
    else:
        if not selected_moodles:
            messages.error(request, "No mentees selected.")
            return redirect("mentee_documents")
        # keep only mentor's mentees
        target_moodles = [m for m in selected_moodles if m in mentees]

    if not target_moodles:
        messages.error(request, "Selected Moodle IDs are not one of your mentees.")
        return redirect("mentee_documents")

    # If no types selected => include all
    ALL_TYPES = {
        "Internship / PBL",
        "Sports / Cultural",
        "Other Event",
        "Course",
        "Publication",
        "Semester Result",
    }
    if selected_types:
        allowed_types = set(t for t in selected_types if t in ALL_TYPES)
    else:
        allowed_types = ALL_TYPES

    files_to_zip = []  # list of tuples: (fs_path, arcname_in_zip)

    def safe_name(s: str) -> str:
        """Sanitize a single filename/folder name (NOT a path)."""
        s = (s or "").strip()
        return "".join(c if c.isalnum() or c in " ._-" else "_" for c in s)

    def safe_path(path: str) -> str:
        """Sanitize a path while keeping folder separators."""
        return "/".join(safe_name(part) for part in (path or "").split("/"))

    def try_add(filefield, arcname):
        if not filefield:
            return
        try:
            fs_path = filefield.path
        except Exception:
            return
        if fs_path and os.path.exists(fs_path):
            files_to_zip.append((fs_path, arcname))

    for moodle in target_moodles:
        mentee_user = mentees[moodle]
        profile = Profile.objects.filter(user=mentee_user).first()
        student_name = (getattr(profile, "student_name", "") or getattr(profile, "name", "") or "Student").strip()

        # Put each mentee into its own folder in the zip
        base_folder = safe_name(f"{moodle}_{student_name}")

        if "Internship / PBL" in allowed_types:
            for item in InternshipPBL.objects.filter(user=mentee_user):
                title = safe_name((item.title or "").strip()[:40])
                try_add(item.certificate, f"{base_folder}/Internship_PBL/{title or 'certificate'}_{os.path.basename(item.certificate.name)}")

        if "Sports / Cultural" in allowed_types:
            for item in SportsCulturalEvent.objects.filter(user=mentee_user):
                ev = safe_name((item.name_of_event or "").strip()[:40])
                try_add(item.certificate, f"{base_folder}/Sports_Cultural/{ev or 'certificate'}_{os.path.basename(item.certificate.name)}")

        if "Other Event" in allowed_types:
            for item in OtherEvent.objects.filter(user=mentee_user):
                ev = safe_name((item.name_of_event or "").strip()[:40])
                try_add(item.certificate, f"{base_folder}/Other_Event/{ev or 'certificate'}_{os.path.basename(item.certificate.name)}")

        if "Course" in allowed_types:
            for item in CertificationCourse.objects.filter(user=mentee_user):
                title = safe_name((item.title or "").strip()[:40])
                try_add(item.certificate, f"{base_folder}/Course/{title or 'certificate'}_{os.path.basename(item.certificate.name)}")

        if "Publication" in allowed_types:
            for item in PaperPublication.objects.filter(user=mentee_user):
                title = safe_name((item.title or "").strip()[:40])
                try_add(item.certificate, f"{base_folder}/Publication/{title or 'certificate'}_{os.path.basename(item.certificate.name)}")

        if "Semester Result" in allowed_types:
            for item in SemesterResult.objects.filter(user=mentee_user):
                sem = safe_name(str(item.semester or ""))
                try_add(item.marksheet, f"{base_folder}/Semester_Result/Sem_{sem}_{os.path.basename(item.marksheet.name)}")

    if not files_to_zip:
        messages.info(request, "No documents available for the selected mentee(s) and filter.")
        return redirect("mentee_documents")

    # Create zip in-memory
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        used = set()
        for fs_path, arcname in files_to_zip:
            arcname = safe_path(arcname)  # IMPORTANT: keep slashes

            candidate = arcname
            base, ext = os.path.splitext(candidate)
            i = 1
            while candidate in used:
                candidate = f"{base}_{i}{ext}"
                i += 1
            used.add(candidate)

            zf.write(fs_path, arcname=candidate)

    buffer.seek(0)

    # filename
    scope_label = "ALL" if download_scope == "all" else "SELECTED"
    zip_filename = f"{scope_label}_mentees_documents.zip"

    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
    return response
#-------------------Download mentee docs logic ends----------------------

#-----------------Download student data logic starts--------------------
# Mapping of all models and their fields
CATEGORY_MAP = {
    "internship": (
        "Internships/PBL",
        InternshipPBL,
        ["title", "academic_year", "semester", "type", "company_name",
         "start_date", "end_date", "no_of_days"]
    ),

    "projects": (
        "Projects",
        Project,
        ["title", "academic_year", "semester", "project_type",
         "guide_name", "link"]
    ),

    "sports": (
        "Sports/Cultural",
        SportsCulturalEvent,
        ["name_of_event", "academic_year", "semester", "type",
         "venue", "level", "prize_won"]
    ),

    "certifications": (
        "Courses/Certifications",
        CertificationCourse,
        ["title", "academic_year", "semester", "domain", "level",
         "start_date", "end_date", "no_of_days"]
    ),

    "other": (
        "Other Achievements",
        OtherEvent,
        ["name_of_event", "academic_year", "semester", "level",
         "details", "prize_won", "amount_won"]
    ),

    "publications": (
        "Paper Publications",
        PaperPublication,
        ["title", "academic_year", "semester", "type", "level",
         "authors", "amount_reimbursed"]
    ),
}

# Category mapping
CATEGORY_MODELS = {
    "internship": InternshipPBL,
    "projects": Project,
    "sports": SportsCulturalEvent,
    "other": OtherEvent,
    "courses": CertificationCourse,
    "publications": PaperPublication,
}


@login_required
def download_student_data(request):
    """
    Mentor-only view to download mentees' data.
    POST params:
      - category (all|internship|projects|sports|other|courses|publications)
      - year (e.g. 2023-24 or ALL)
      - branch (ALL or branch code)
      - export (excel|pdf|zip)
    """
    # --- basic context & permission check ---
    mentor = get_object_or_404(Mentor, user=request.user)
    mentee_mappings = MentorMentee.objects.filter(mentor=mentor).select_related("mentee__user")
    user_list = [m.mentee.user for m in mentee_mappings]

    profiles_qs = Profile.objects.filter(user__in=user_list)

    context = {
        "mentees": profiles_qs,
        "categories": [
            ("all", "ALL Categories (Combined)"),
            ("internship", "Internship / PBL"),
            ("projects", "Projects"),
            ("sports", "Sports & Cultural"),
            ("other", "Other Achievements / Hackathons"),
            ("courses", "Courses / Certifications"),
            ("publications", "Paper Publications"),
        ],
        "years": [f"{y}-{str(y+1)[-2:]}" for y in range(2017, 2027)],
        "branches": ["ALL", "IT", "CSE", "AIML", "DS", "MECH", "CIVIL"],
    }

    if request.method == "GET":
        return render(request, "mentor/download_student_data.html", context)

    # --- Read POST inputs ---
    category = request.POST.get("category", "all")
    academic_year = request.POST.get("year", "ALL")
    branch = request.POST.get("branch", "ALL")
    export_type = request.POST.get("export", "excel")  # excel|pdf|zip

    # Branch filter on profiles (applies globally)
    if branch != "ALL":
        profiles_qs = profiles_qs.filter(branch=branch)

    # Helper: get profile for a user (cached by dict for speed)
    profile_map = {p.user_id: p for p in profiles_qs}
    def profile_for_user(user):
        return profile_map.get(user.pk)

    # --- Collect rows for a model safely ---
    def collect_entries(model):
        qs = model.objects.filter(user__in=user_list)

        # Only filter by academic_year if model actually has this field and user selected a specific year
        if academic_year and academic_year != "ALL":
            if "academic_year" in [f.name for f in model._meta.fields]:
                qs = qs.filter(academic_year=academic_year)

        # If branch selected, limit to users whose Profile branch matches
        if branch != "ALL":
            valid_user_ids = list(profiles_qs.values_list("user_id", flat=True))
            qs = qs.filter(user__in=valid_user_ids)

        rows = []
        for obj in qs.select_related("user"):
            prof = profile_for_user(obj.user)
            base = {
                "moodle_id": prof.moodle_id if prof else getattr(obj.user, "username", ""),
                "student_name": prof.student_name if prof else getattr(obj.user, "username", ""),
                "email": getattr(obj.user, "email", "") or "",
                "branch": prof.branch if prof else "",
                "_model": model.__name__,
                "_pk": obj.pk,
            }

            # iterate model fields and capture values (exclude id,user)
            for f in model._meta.fields:
                if f.name in ("id", "user", "certificate", "certificates"):
                    continue
                val = getattr(obj, f.name)

                # datetime tz fix
                if hasattr(val, "tzinfo") and val.tzinfo:
                    val = val.replace(tzinfo=None)

                # FileField -> output URL (if present) else file name
                from django.db.models.fields.files import FieldFile
                if isinstance(val, FieldFile):
                    if val and getattr(val, "name", ""):
                        try:
                            # prefer absolute url if available
                            url = val.url
                            full_url = request.build_absolute_uri(url)
                            val = full_url
                        except Exception:
                            val = val.name or ""
                    else:
                        val = ""

                # coerce to string for safe storage in collected rows
                if isinstance(val, (list, dict)):
                    # JSON-serialize complex types
                    try:
                        val = json.dumps(val, default=str)
                    except:
                        val = str(val)
                elif val is None:
                    val = ""

                base[f.name] = val
            base.pop("certificate", None)
            base.pop("certificates", None)
            rows.append(base)
        return rows

    # --- Build collected dict for chosen categories ---
    collected = {}
    for key, model in CATEGORY_MODELS.items():
        if category != "all" and category != key:
            continue
        collected[key] = collect_entries(model)

    # --- Helpers used by Excel builder ---
    def auto_adjust(ws):
        """Auto adjust column widths"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    v = str(cell.value) if cell.value is not None else ""
                    if len(v) > max_len:
                        max_len = len(v)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    # -------------------------
    # Build Excel workbook
    # -------------------------
    def build_excel_workbook():
        wb = openpyxl.Workbook()
        cover = wb.active
        cover.title = "Cover"
        cover["A1"] = "APSIT - Student Data Export"
        cover["A1"].font = Font(size=18, bold=True)
        cover["A3"] = "Generated by:"
        cover["B3"] = request.user.get_full_name() or request.user.username
        cover["A4"] = "Generated on:"
        cover["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cover["A5"] = "Selected Category:"
        cover["B5"] = category
        cover["A6"] = "Academic Year:"
        cover["B6"] = academic_year
        cover["A7"] = "Branch:"
        cover["B7"] = branch

        # Summary sheet
        summary = wb.create_sheet("Summary")
        summary["A1"] = "Summary Analytics"
        summary["A1"].font = Font(size=14, bold=True)

        total_mentees = profiles_qs.count()
        total_docs = sum(len(collected.get(k, [])) for k in collected)
        summary_data = [
            ("Total Mentees", total_mentees),
            ("Total Records (selected categories)", total_docs),
        ]

        # Add counts for each category + subcounts
        if "internship" in collected:
            ints = collected["internship"]
            summary_data.append(("Total Internships/PBL Records", len(ints)))
            type_counter = Counter([r.get("type") for r in ints if r.get("type")])
            for t, c in type_counter.items():
                summary_data.append((f"Internship Type: {t}", c))

        if "projects" in collected:
            summary_data.append(("Total Projects Records", len(collected["projects"])))

        if "sports" in collected:
            summary_data.append(("Total Sports/Cultural Records", len(collected["sports"])))
            level_counts = Counter([r.get("level") for r in collected["sports"] if r.get("level")])
            for lvl, c in level_counts.items():
                summary_data.append((f"Sports Level: {lvl}", c))

        if "other" in collected:
            summary_data.append(("Total Other Achievements Records", len(collected["other"])))
            prize_counts = Counter([r.get("prize_won") for r in collected["other"] if r.get("prize_won")])
            for p, c in prize_counts.items():
                summary_data.append((f"Prize Won ({p})", c))

        if "courses" in collected:
            summary_data.append(("Total Certifications Records", len(collected["courses"])))
            auth_counts = Counter([r.get("certifying_authority") for r in collected["courses"] if r.get("certifying_authority")])
            for a, c in auth_counts.items():
                summary_data.append((f"Authority: {a}", c))

        if "publications" in collected:
            summary_data.append(("Total Publications Records", len(collected["publications"])))
            pub_counts = Counter([r.get("type") for r in collected["publications"] if r.get("type")])
            for t, c in pub_counts.items():
                summary_data.append((f"Publication Type: {t}", c))

        # write summary_data
        r = 3
        for k, v in summary_data:
            summary[f"A{r}"] = k
            summary[f"B{r}"] = v
            r += 1

        # Branch x Category matrix
        branches = ["IT", "CSE", "AIML", "DS", "MECH", "CIVIL"]
        categories = list(CATEGORY_MODELS.keys())
        matrix_start_row = r + 2
        summary.cell(row=matrix_start_row, column=1, value="Branch-wise Total Records").font = Font(size=12, bold=True)

        header_row_idx = matrix_start_row + 1
        # header
        for ci, h in enumerate(["Branch"] + [c.title() for c in categories], start=1):
            summary.cell(row=header_row_idx, column=ci, value=h)

        # data rows
        for bi, br in enumerate(branches, start=header_row_idx + 1):
            summary.cell(row=bi, column=1, value=br)
            for ci, cat in enumerate(categories, start=2):
                cnt = sum(1 for rdict in collected.get(cat, []) if rdict.get("branch") == br)
                summary.cell(row=bi, column=ci, value=cnt)

        auto_adjust(summary)

        # Create per-category sheets with table & autofilter
        for key, rows in collected.items():
            sheet_name = key.upper()[:31]
            sheet = wb.create_sheet(sheet_name)
            if not rows:
                sheet.append(["No records found"])
                continue

            # columns from first row (stable ordering)
            first = rows[0]
            cols = [c for c in first.keys() if not c.startswith("_")]
            sheet.append(cols)

            for rdict in rows:
                row = [rdict.get(c, "") for c in cols]
                sheet.append(row)

            # format table & autofilter
            sheet.freeze_panes = sheet["A2"]
            table_ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
            table = Table(displayName=f"{key}_table", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            sheet.add_table(table)

            auto_adjust(sheet)
        return wb

    # -------------------------
    # Build PDF bytes (ReportLab)
    # -------------------------
    def add_footer(canvas, doc):
        canvas.setFont("Helvetica", 9)

        # --- Page number ---
        page_num_text = f"Page {canvas.getPageNumber()}"

        # --- Mentor name ---
        mentor_name = (
                doc._request.user.mentor.name
                or doc._request.user.get_full_name()
                or doc._request.user.username
        )
        mentor_text = f"Mentor: {mentor_name}"

        # --- Bottom Y position ---
        y = 1 * cm

        # Page number (center)
        canvas.drawCentredString(A4[0] / 2.0, y, page_num_text)

        # Left text (mentor)
        canvas.drawString(1.5 * cm, y, mentor_text)

    def build_pdf_bytes():
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
        # inject request for footer
        doc._request = request
        story = []
        styles = getSampleStyleSheet()
        heading_style = styles["Heading1"]
        heading_style.fontSize = 16

        # logo
        logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")
        if os.path.exists(logo_path):
            try:
                img = RLImage(logo_path, width=2.5 * cm, height=2 * cm)
            except Exception:
                img = None

        # header
        header_cells = []
        if img:
            header_cells.append(img)
        header_cells.append(Paragraph("<b>APSIT - Student Data</b>", heading_style))
        header_table = RLTable([[header_cells[0] if img else "", header_cells[-1]]], colWidths=[2.4 * cm, 14 * cm])
        story.append(header_table)
        story.append(Spacer(1, 8))
        meta = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Mentor: {request.user.mentor.name or request.user.get_full_name() or request.user.username}"
        story.append(Paragraph(meta, styles["Normal"]))
        story.append(Spacer(1, 12))

        # each category
        small_para_style = ParagraphStyle('small', fontSize=8, leading=10)
        for key, rows in collected.items():
            title = key.replace("_", " ").title()
            story.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))
            story.append(Spacer(1, 6))
            if not rows:
                story.append(Paragraph("<i>No records found.</i>", styles["BodyText"]))
                story.append(PageBreak())
                continue

            # columns
            first = rows[0]
            cols = [c for c in first.keys() if not c.startswith("_")]
            data = [cols]
            for rdict in rows:
                data.append([str(rdict.get(c, "") or "") for c in cols])

            # convert to Paragraphs so cell text wraps
            wrapped = []
            for row in data:
                wrapped_row = [Paragraph(cell.replace("\n", "<br/>"), small_para_style) for cell in row]
                wrapped.append(wrapped_row)

            page_width = A4[0] - (doc.leftMargin + doc.rightMargin)
            col_count = max(1, len(cols))
            col_width = page_width / col_count
            col_widths = [col_width] * col_count

            tbl = RLTable(wrapped, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(RLTableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3f5cbf")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(tbl)
            story.append(PageBreak())

        doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0)
        return buffer

    # -------------------------
    # Build CSV & JSON for zip
    # -------------------------
    def build_csv_and_json():
        files = {}
        import csv
        from io import StringIO

        for key, rows in collected.items():
            if not rows:
                continue

            # -----------------------
            # CSV (StringIO â†’ encode)
            # -----------------------
            csv_io = StringIO()
            writer = csv.writer(csv_io)

            first = rows[0]
            cols = [c for c in first.keys() if not c.startswith("_")]
            writer.writerow(cols)

            for r in rows:
                writer.writerow([r.get(c, "") for c in cols])

            files[f"{key}.csv"] = csv_io.getvalue().encode("utf-8")

            # -----------------------
            # JSON
            #------------------------
            files[f"{key}.json"] = json.dumps(rows, default=str, indent=2).encode("utf-8")

        return files

    # -------------------------
    # Deliver based on export_type
    # -------------------------
    if export_type == "pdf":
        request.session["export_done"] = True
        pdf_buf = build_pdf_bytes()
        return FileResponse(pdf_buf, as_attachment=True, filename=f"{category}_AY-{academic_year}_{branch}-Dept_student_data.pdf")

    if export_type == "excel":
        request.session["export_done"] = True
        wb = build_excel_workbook()
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return FileResponse(out, as_attachment=True, filename=f"Excel_{category}_AY-{academic_year}_{branch}-Dept_student_data.xlsx", content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if export_type == "zip":
        request.session["export_done"] = True
        wb = build_excel_workbook()
        excel_buf = BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        pdf_buf = build_pdf_bytes()
        csvs = build_csv_and_json()

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"Excel_{category}_AY-{academic_year}_{branch}-Dept_student_data.xlsx", excel_buf.getvalue())
            zf.writestr(f"{category}_AY-{academic_year}_{branch}-Dept_student_data.pdf", pdf_buf.getvalue())
            for name, data in csvs.items():
                zf.writestr(name, data)
        zip_buffer.seek(0)
        return FileResponse(zip_buffer, as_attachment=True, filename=f"{category}_AY-{academic_year}_{branch}-Dept_student_data_bundle.zip")

    # fallback
    return render(request, "mentor/download_student_data.html", context)
#-----------------Download student data logic ends--------------------


@login_required
def clear_export_flag(request):
    request.session.pop("export_done", None)
    return HttpResponse("OK")


#--------------------Messages page logic starts----------------------
#--------------------Inbox requests and chatting logic-----------------------
@method_decorator(login_required, name='dispatch')
class MessageView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """controls message view"""

    template_name = 'mentor/messages-module1.html'
    model = Msg

    def test_func(self):
        return self.request.user.is_mentor

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['count'] = Msg.objects.filter(receipient=self.request.user).filter(is_approved=False).count()
        context['count1'] = Msg.objects.filter(receipient=self.request.user).filter(is_approved=True).count()
        context['count2'] = Conversation.objects.filter(sender=self.request.user).count()
        return context

    def get_queryset(self):
        return self.model.objects.filter(receipient=self.request.user)

@method_decorator(login_required, name='dispatch')
class MessageCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Creates new message"""

    fields = ('receipient', 'msg_content')
    model = Msg
    template_name = 'mentor/messagecreate1.html'

    def test_func(self):
        return self.request.user.is_mentor

    def form_valid(self, form):
        form.instance.sender = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('list1')

@method_decorator(login_required, name='dispatch')
class MessageListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """List sent Messages"""

    model = Msg
    template_name = 'mentor/listmessages1.html'
    context_object_name = 'sentmesso'

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        return self.model.objects.filter(sender=self.request.user)

@method_decorator(login_required, name='dispatch')
class SentDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """details the message sent"""

    model = Msg
    context_object_name = 'messo'
    template_name = 'mentor/sent1.html'

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        return self.model.objects.filter(sender=self.request.user)

@method_decorator(login_required, name='dispatch')
class SentMessageDelete(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Deletes sent messages"""

    model = Msg
    success_url = reverse_lazy("list1")
    template_name = 'mentor/sentmessage_delete1.html'

    def test_func(self):
        return self.request.user.is_mentor

@method_decorator(login_required, name='dispatch')
class InboxView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """ Lists messages in inbox view"""

    def test_func(self):
        return self.request.user.is_mentor

    model = Msg
    context_object_name = 'inbox'
    template_name = 'mentor/inbox1.html'
    paginate_by = 10

    def get_queryset(self):
        return self.model.objects.filter(receipient=self.request.user).filter(is_approved=False)

@method_decorator(login_required, name='dispatch')
class InboxDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Inbox Detailed view"""
    model = Msg
    context_object_name = 'messo'
    template_name = 'mentor/inboxview1.html'

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        return self.model.objects.filter(receipient=self.request.user)


@login_required
def reply_message(request, pk):
    """Replies, approves, comments on messages"""

    if not request.user.is_mentor:
        return redirect('home')

    reply = get_object_or_404(Msg, pk=pk, receipient=request.user)

    if request.method == 'POST':

        form = ReplyForm(request.POST)

        if form.is_valid():
            reply.is_approved = form.cleaned_data['is_approved']
            reply.comment = form.cleaned_data['comment']
            reply.save()
            # ðŸ”” Notify mentee
            Notification.objects.create(
                user=reply.sender,  # sender is the mentee
                message="Your request has been accepted by your mentor."
            )
            return redirect('approved1')

    else:
        form = ReplyForm

    context = {
        'messo': reply,
        'reply': reply,
        'form': form,
    }

    return render(request, 'mentor/comment.html', context)

@method_decorator(login_required, name='dispatch')
class Approved(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """view list of approved messeges from mentors"""

    def test_func(self):
        return self.request.user.is_mentor

    model = Msg.objects.filter(is_approved=True).order_by('-date_approved')

    template_name = 'mentor/approved.html'

    context_object_name = 'messo'

    paginate_by = 10

    def get_queryset(self):
        return self.model.filter(receipient=self.request.user)


@method_decorator(login_required, name='dispatch')
class ProfileDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """view details of a user in the profile"""

    model = Msg
    context_object_name = 'msg'
    template_name = 'mentor/profile_detail1.html'

    def test_func(self):
        return self.request.user.is_mentor


@method_decorator(login_required, name='dispatch')
class ConversationCreateView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView):
    """Create Conversation"""

    fields = ('conversation',)
    model = Conversation
    template_name = 'mentor/chat.html'
    context_object_name = 'conversation'
    success_message = 'Your Conversation Has been Created!'

    def test_func(self):
        return self.request.user.is_mentor

    def form_valid(self, form):
        form.instance.sender = self.request.user
        form.instance.receipient = User.objects.get(pk=self.kwargs['pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('conv1')


@method_decorator(login_required, name='dispatch')
class ConversationListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """List all chat conversation by a user"""

    model = Conversation
    template_name = 'mentor/list-converations.html'
    context_object_name = 'conversation'
    paginate_by = 10

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        return self.model.objects.filter(sender=self.request.user)


@method_decorator(login_required, name='dispatch')
class ReplyCreateView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView):
    """Replies by a user"""

    fields = ('reply',)
    model = Reply
    template_name = 'mentor/conversation.html'
    success_message = 'You have replied!'

    def test_func(self):
        return self.request.user.is_mentor

    def form_valid(self, form):
        form.instance.sender = self.request.user
        form.instance.conversation = Conversation.objects.get(pk=self.kwargs['pk'])
        return super().form_valid(form)

    def get_success_url(self):
        conversation = self.object.conversation
        return reverse_lazy('conv-reply', kwargs={'pk': self.object.conversation_id})

    def get_queryset(self):
        return self.model.objects.filter(sender=self.request.user)


@method_decorator(login_required, name='dispatch')
class ConversationDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Conversation
    template_name = 'mentor/conversation1.html'
    context_object_name = 'conv'

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        from django.db.models import Q
        return Conversation.objects.filter(
            Q(sender=self.request.user) | Q(receipient=self.request.user)
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ChatReplyForm()
        context["is_mentor_view"] = True
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = ChatReplyForm(request.POST, request.FILES)

        if form.is_valid():
            reply = Reply()
            reply.reply = request.POST.get("reply")
            reply.file = request.FILES.get("file")
            reply.sender = request.user
            reply.conversation = self.object
            reply.replied_at = now()
            reply.save()

            # Identify mentee (the other participant)
            mentee_user = (
                self.object.receipient
                if self.object.sender == request.user
                else self.object.sender
            )

            Notification.objects.create(
                user=mentee_user,
                message="Your mentor has replied to your conversation."
            )

            return redirect('conv-reply', pk=self.object.pk)

        context = self.get_context_data(object=self.object)
        context['form'] = form
        return self.render_to_response(context)



@method_decorator(login_required, name='dispatch')
class ConversationDeleteView(SuccessMessageMixin, DeleteView):
    """delete view Chat"""

    model = Reply
    template_name = 'mentor/chat-confirm-delete.html'
    success_message = 'Your message has been deleted!'

    # success_url = reverse_lazy('conv1')

    def get_success_url(self):
        conversation = self.object.conversation
        return reverse_lazy('conv-reply', kwargs={'pk': self.object.conversation_id})


@method_decorator(login_required, name='dispatch')
class Conversation2DeleteView(DeleteView):
    """delete view Coversation"""

    model = Conversation
    template_name = 'mentor/conversation-confirm-delete.html'

    # success_url = reverse_lazy('conv1')

    def get_success_url(self):
        return reverse_lazy('conv1')


@method_decorator(login_required, name='dispatch')
class MentorMeetingListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Meeting
    template_name = 'mentor/vc.html'
    context_object_name = 'meetings'
    paginate_by = 10

    def test_func(self):
        return self.request.user.is_mentor

    def get_queryset(self):
        qs = Meeting.objects.filter(
            mentor__user=self.request.user
        ).select_related(
            "mentee__user", "mentee__user__profile"
        ).order_by('-appointment_date', '-time_slot')

        # ðŸ” FILTERS
        moodle_id = self.request.GET.get("moodle_id")
        name = self.request.GET.get("name")
        date = self.request.GET.get("date")

        if moodle_id:
            qs = qs.filter(
                mentee__user__profile__moodle_id__icontains=moodle_id
            )

        if name:
            qs = qs.filter(
                mentee__user__profile__student_name__icontains=name
            )

        if date:
            qs = qs.filter(appointment_date=date)

        # âœ… Auto-complete past meetings
        now = timezone.localtime(timezone.now())
        for meeting in qs:
            if meeting.status == 'scheduled' and now > meeting.meeting_end_datetime:
                meeting.status = 'completed'
                meeting.save(update_fields=['status'])

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["now"] = timezone.localtime(timezone.now())

        # preserve filter values
        context["selected_moodle_id"] = self.request.GET.get("moodle_id", "")
        context["selected_name"] = self.request.GET.get("name", "")
        context["selected_date"] = self.request.GET.get("date", "")

        return context


@login_required
def join_meeting(request, meeting_id):
    meeting = get_object_or_404(Meeting, id=meeting_id)

    if request.user not in [meeting.mentor.user, meeting.mentee.user]:
        return redirect('home')

    now = timezone.localtime(timezone.now())

    # âŒ Do not allow joining expired meetings
    if now > meeting.meeting_end_datetime:
        if meeting.status != 'completed':
            meeting.status = 'completed'
            meeting.save(update_fields=['status'])
        return redirect('vc1')

    # âœ… DO NOT mark completed here
    return redirect(f"https://meet.jit.si/{meeting.video_room_name}")


@login_required
def mentor_queries(request):
    try:
        mentor = Mentor.objects.get(user=request.user)
    except Mentor.DoesNotExist:
        return redirect('mentee_queries')

    queries = Query.objects.filter(mentor=mentor).select_related(
        "mentee__user", "mentee__user__profile"
    ).order_by('-created_at')

    # ðŸ” FILTERS (GET)
    moodle_id = request.GET.get("moodle_id")
    name = request.GET.get("name")
    severity = request.GET.get("severity")

    if moodle_id:
        queries = queries.filter(mentee__user__profile__moodle_id__icontains=moodle_id)

    if name:
        queries = queries.filter(
            mentee__user__profile__student_name__icontains=name
        )

    if severity:
        queries = queries.filter(severity=severity)

    # âœ… BULK MARK AS DONE (POST)
    if request.method == "POST" and "mark_all_done" in request.POST:
        pending = queries.filter(status="pending")

        for q in pending:
            q.status = "resolved"
            q.save()

            # ðŸ”” notify mentee
            Notification.objects.create(
                user=q.mentee.user,
                message="Your query has been resolved by your mentor."
            )

        messages.success(request, f"{pending.count()} queries marked as resolved.")
        return redirect("mentor_queries")

    # ðŸ“„ PAGINATION
    paginator = Paginator(queries, 8)  # 8 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "mentor/mentor_queries.html", {
        "queries": page_obj,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "selected_moodle_id": moodle_id or "",
        "selected_name": name or "",
        "selected_severity": severity or "",
    })


@login_required
def mark_as_done(request, query_id):
    query = get_object_or_404(Query, id=query_id)
    query.status = "resolved"
    query.save()

    # ðŸ”” CREATE NOTIFICATION FOR MENTEE
    Notification.objects.create(
        user=query.mentee.user,
        message=f"Your query has been resolved by Mentor {query.mentor.name}."
    )

    messages.success(request, f"Query from {query.mentee} - {query.mentee.profile.student_name} has been resolved.")
    return redirect(reverse('mentor_queries'))
#--------------------Messages page logic ends----------------------

#-----------------forget password logic starts---------------------
def send_forget_password_email(email, token):
    try:
        subject = 'Reset Your Password'
        reset_link = f'http://127.0.0.1:8000/change-pass/{token}/'
        message = f"""
        Hi,

        You requested a password reset. Click the link below to reset your password:

        {reset_link}

        If you did not request this, please ignore this email.

        Regards,
        APSIT
        """
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [email]

        send_mail(subject, message.strip(), email_from, recipient_list)
        print("âœ… Reset email sent successfully.")
        return True

    except Exception as e:
        print(f"âŒ Email send failed: {e}")
        return False


def ForgetPass(request):
    try:
        if request.method == 'POST':
            username = request.POST.get('username')
            print(f"Submitted username: {username}")

            # Find the user by username
            user_obj = User.objects.filter(username=username).first()
            if not user_obj:
                messages.error(request, 'No user found with this username.')
                return render(request, 'mentor/forget-pass.html')

            # Get the user's profile
            profile_obj = Profile.objects.filter(user=user_obj).first()
            if not profile_obj:
                messages.error(request, 'Profile not found for this user.')
                return render(request, 'mentor/forget-pass.html')

            # Create token and save
            token = str(uuid.uuid4())
            profile_obj.forget_password_token = token
            profile_obj.save()

            # Send email
            send_forget_password_email(user_obj.email, token)

            messages.success(request, 'Password reset link has been sent to your email.')
            return render(request, 'mentor/forget-pass.html')

    except Exception as e:
        print("Error in ForgetPassword:")
        traceback.print_exc()
        messages.error(request, 'Something went wrong. Please try again.')

    return render(request, 'mentor/forget-pass.html')

def ChangePass(request, token):
    try:
        profile_obj = Profile.objects.filter(forget_password_token=token).first()

        if not profile_obj:
            messages.error(request, "Invalid or expired token.")
            return redirect('login1')

        context = {
            'user_id': profile_obj.user.id,
            'token': token
        }

        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('reconfirm_password')

            if new_password != confirm_password:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'mentor/change-pass.html', context)

            user = profile_obj.user
            user.set_password(new_password)
            user.save()

            profile_obj.forget_password_token = None
            profile_obj.save()

            messages.success(request, 'Password updated successfully.')
            return redirect('login1')     # Mentor login

        return render(request, 'mentor/change-pass.html', context)

    except Exception as e:
        print("Error in ChangePass:", e)
        messages.error(request, "Something went wrong. Try again.")
        return redirect('forget_pass')
#-----------------forget password logic ends---------------------

#----------------Mentor-Mentee interaction page logic starts-----------------
COMMON_AGENDA_POINTS = [
    "Academic performance review",
    "75% Attendance compulsory",
    "Internship / Project progress",
    "Career guidance & higher studies",
    "Personal or behavioural concerns",
]

def roman_to_int(r: str):
    if not r:
        return None
    r = str(r).strip().upper()
    roman_map = {"I": 1, "V": 5, "X": 10}

    total = 0
    prev = 0
    for ch in reversed(r):
        val = roman_map.get(ch)
        if not val:
            return None
        if val < prev:
            total -= val
        else:
            total += val
            prev = val
    return total

def sem_to_class(sem: str) -> str:
    n = roman_to_int(sem)
    if not n:
        return ""
    if n in (1, 2): return "FE"
    if n in (3, 4): return "SE"
    if n in (5, 6): return "TE"
    if n in (7, 8): return "BE"
    return ""

@login_required
def mentor_mentee_interactions(request):

    mentor = get_object_or_404(Mentor, user=request.user)

    # âœ… Get mentees under this mentor
    mappings = MentorMentee.objects.filter(mentor=mentor).select_related("mentee__user")
    mentee_users = [m.mentee.user for m in mappings]
    mentee_profiles = Profile.objects.filter(user__in=mentee_users)

    # âœ… Filtering
    semester_filter = request.GET.get("semester")
    date_filter = request.GET.get("date")
    class_filter = request.GET.get("class_year")

    interactions = MentorMenteeInteraction.objects.filter(mentor=request.user)

    if semester_filter:
        interactions = interactions.filter(semester=semester_filter)

    if date_filter:
        interactions = interactions.filter(date=date_filter)

    if class_filter:
        interactions = interactions.filter(class_year__icontains=class_filter)

    # âœ… CREATE + UPDATE INTERACTION (SAME FORM)
    if request.method == "POST":
        interaction_id = request.POST.get("interaction_id")

        date = request.POST.get("date")
        mentee_ids = request.POST.getlist("mentees")
        checked_points = request.POST.getlist("agenda_points")
        extra_text = request.POST.get("agenda_extra", "").strip()

        # âœ… FIXED: Auto semester from SELECTED mentees only
        semester = ""
        class_year = ""
        selected_profiles = Profile.objects.filter(user__id__in=mentee_ids)
        if selected_profiles.exists():
            # âœ… semesters (Roman numerals)
            semesters = list(
                selected_profiles.values_list("semester", flat=True).distinct()
            )
            semesters = [s for s in semesters if s]  # remove None/empty

            semester = ", ".join(semesters)

            # âœ… derive FE/SE/TE/BE from Roman numerals
            classes = sorted({sem_to_class(s) for s in semesters if sem_to_class(s)})
            class_year = ", ".join(classes)  # e.g. "FE" or "FE, SE"

        # âœ… Build agenda safely
        agenda_parts = checked_points[:]
        if extra_text:
            agenda_parts.append(extra_text)
        agenda_text = "; ".join(agenda_parts)

        # âœ… UPDATE EXISTING
        if interaction_id:
            interaction = get_object_or_404(
                MentorMenteeInteraction, id=interaction_id, mentor=request.user
            )
            interaction.date = date
            interaction.semester = semester
            interaction.class_year = class_year
            interaction.agenda = agenda_text
            # âœ… since agenda changed, invalidate old summary unless you want to keep it
            interaction.ai_summary = None
            interaction.ai_summary_generated = False
            interaction.save()
            interaction.mentees.set(User.objects.filter(id__in=mentee_ids))

            messages.success(request, "Interaction updated successfully.")

        # âœ… CREATE NEW
        else:
            interaction = MentorMenteeInteraction.objects.create(
                mentor=request.user,
                date=date,
                semester=semester,
                class_year=class_year,
                agenda=agenda_text,
                ai_summary=None,  # âœ… ensure empty
                ai_summary_generated=False,  # âœ… manual only
            )
            interaction.mentees.set(User.objects.filter(id__in=mentee_ids))

            messages.success(request, "Interaction saved successfully.")

        return redirect("mentor_interaction")

    # âœ… ATTENDANCE REPORT
    attendance = []
    total_interactions = interactions.count()

    for p in mentee_profiles:
        present = interactions.filter(mentees=p.user).count()
        percent = round((present / total_interactions) * 100, 2) if total_interactions else 0

        attendance.append({
            "name": p.student_name,
            "semester": p.semester,
            "year": p.year,
            "branch": p.branch,
            "moodle": p.moodle_id,
            "present": present,
            "total": total_interactions,
            "percent": percent
        })
    #monthly attendance
    monthly_raw = (
        interactions
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )

    monthly_attendance = []

    for row in monthly_raw:
        month_name = row["month"].strftime("%b %Y")
        percent = round((row["total"] / total_interactions) * 100, 2) if total_interactions else 0
        monthly_attendance.append({
            "month": month_name,
            "percent": percent
        })

    # âœ… Pagination
    paginator = Paginator(interactions.order_by("-date"), 8)  # 8 rows per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "mentor": mentor,
        "mentees": mentee_profiles,
        "interactions": page_obj,
        "page_obj": page_obj,
        "common_agenda": COMMON_AGENDA_POINTS,
        "attendance": attendance,
        "semesters": [s for s in SEMESTER],
        "monthly_attendance": monthly_attendance,
    }

    return render(request, "mentor/mentor_mentee_interactions.html", context)


@login_required
def delete_interaction(request, pk):
    interaction = get_object_or_404(MentorMenteeInteraction, pk=pk, mentor=request.user)
    interaction.delete()
    messages.success(request, "Interaction deleted.")
    return redirect("mentor_interaction")


@login_required
@require_POST
def regenerate_ai_summary(request, pk):
    interaction = get_object_or_404(MentorMenteeInteraction, id=pk, mentor=request.user)

    # âœ… Call your AI generator here
    summary = generate_ai_summary(interaction.agenda)

    interaction.ai_summary = summary
    interaction.ai_summary_generated = True  # âœ… mark manual generation
    interaction.save(update_fields=["ai_summary", "ai_summary_generated"])

    return JsonResponse({
        "success": True,
        "summary": interaction.ai_summary
    })


@require_POST
@login_required
def remove_ai_summary(request, pk):
    interaction = get_object_or_404(
        MentorMenteeInteraction,
        id=pk,
        mentor=request.user
    )

    # âœ… Save for undo
    interaction.last_ai_summary = interaction.ai_summary

    # âœ… Clear active summary
    interaction.ai_summary = None
    interaction.ai_summary_generated = False

    interaction.save(update_fields=[
        "ai_summary",
        "ai_summary_generated",
        "last_ai_summary"
    ])

    return JsonResponse({"success": True})


@require_POST
@login_required
def undo_ai_summary(request, pk):
    interaction = get_object_or_404(
        MentorMenteeInteraction,
        id=pk,
        mentor=request.user
    )

    if not interaction.last_ai_summary:
        return JsonResponse({"success": False, "error": "Nothing to restore"})

    interaction.ai_summary = interaction.last_ai_summary
    interaction.ai_summary_generated = True
    interaction.last_ai_summary = None

    interaction.save(update_fields=[
        "ai_summary",
        "ai_summary_generated",
        "last_ai_summary"
    ])

    return JsonResponse({
        "success": True,
        "summary": interaction.ai_summary
    })


@login_required
def export_interactions(request, export_type):
    mentor = get_object_or_404(Mentor, user=request.user)
    qs = MentorMenteeInteraction.objects.filter(mentor=request.user).prefetch_related("mentees__profile")

    # âœ… EXCEL
    if export_type == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Semester", "Year", "Agenda", "AI Summary", "Mentees"])

        for i in qs:
            mentees = ", ".join([
                f"{u.profile.student_name} ({u.profile.moodle_id})"
                for u in i.mentees.all()
            ])

            ws.append([
                i.date.strftime("%d-%m-%Y"),
                i.semester,
                i.class_year,
                i.agenda,
                i.ai_summary if getattr(i, "ai_summary_generated", False) else "Not Generated",
                mentees
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename="Mentor_Mentee_interactions.xlsx")

    # âœ… PDF EXPORT (FIXED FORMATTING)
    buffer = BytesIO()

    # âœ… Convert Agenda & AI Summary into Bullet List
    def bulletize(text):
        if not text:
            return ""

        text = text.strip()

        # âœ… REMOVE any existing <link> tags (prevents duplication)
        text = re.sub(r"<link[^>]*>|</link>", "", text)

        # âœ… 1. Protect URLs with tokens
        url_pattern = r"(https?://[^\s]+)"
        urls = re.findall(url_pattern, text)

        url_map = {}
        for i, url in enumerate(urls):
            token = f"[[URL_{i}]]"
            url_map[token] = url
            text = text.replace(url, token)

        # âœ… 2. Normalize separators
        text = text.replace("â€¢", ".")
        text = text.replace("\n", ".")
        text = text.replace(";", ".")

        # âœ… 3. Split into sentences safely
        sentences = re.split(r"\.(?=\s|$)", text)

        bullets = []

        for s in sentences:
            s = s.strip()
            if not s:
                continue

            # âœ… 4. Restore each URL as BLUE + UNDERLINED clickable link
            for token, url in url_map.items():
                s = s.replace(
                    token,
                    f'<font color="blue"><u><link href="{url}">{url}</link></u></font>'
                )

            # âœ… 5. Escape only normal text
            s = escape(s, quote=False)

            # âœ… Un-escape the tags we intentionally inserted
            s = s.replace("&lt;font", "<font").replace("&lt;/font&gt;", "</font>")
            s = s.replace("&lt;u&gt;", "<u>").replace("&lt;/u&gt;", "</u>")
            s = s.replace("&lt;link", "<link").replace("&lt;/link&gt;", "</link>")
            s = s.replace("&gt;", ">")

            bullets.append("â€¢ " + s)

        return "<br/>".join(bullets)

    # âœ… HIGH-RES PDF (Better Printing)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=110,
        bottomMargin=90
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="BlueLink",
        parent=styles["Normal"],
        textColor=colors.blue,
        underline=True
    ))
    story = []

    # âœ… EXPORT DATE (TODAY)
    export_date = datetime.now().strftime("%d-%m-%Y %I:%M %p")

    # âœ… GET UNIQUE DEPARTMENTS & SEMESTERS FROM MENTEES
    mentee_departments = set()
    mentee_semesters = set()

    for i in qs:
        for u in i.mentees.all():
            if hasattr(u, "profile"):
                if u.profile.branch:
                    mentee_departments.add(u.profile.branch)
                if u.profile.semester:
                    mentee_semesters.add(u.profile.semester)

    department = ", ".join(sorted(mentee_departments)) if mentee_departments else "Not Available"
    semester_filter = ", ".join(sorted(mentee_semesters)) if mentee_semesters else "Not Available"

    # âœ… HEADER DETAILS
    header_info = f"""
    <b>Department:</b> {department} &nbsp;&nbsp;&nbsp;&nbsp;
    <b>Semester:</b> {semester_filter} <br/>
    <b>Generated on:</b> {export_date}
    """

    # âœ… TABLE DATA
    data = [["Sr. No.", "Date", "Sem", "Year", "Agenda", "AI Summary", "Mentees Present"]]

    for idx, i in enumerate(qs, start=1):
        mentee_lines = []

        for u in i.mentees.all():
            if hasattr(u, "profile"):
                name = u.profile.student_name
                dept = u.profile.branch if u.profile.branch else "N/A"
                moodle = u.profile.moodle_id if u.profile.moodle_id else "N/A"

                mentee_lines.append(f"{name} ({dept}, {moodle})")

        mentees_formatted = "<br/>".join(mentee_lines)

        agenda_paragraph = Paragraph(bulletize(i.agenda), styles["Normal"])
        ai_summary_paragraph = Paragraph(
            bulletize(i.ai_summary) if getattr(i, "ai_summary_generated", False) else bulletize("Not Generated"),
            styles["Normal"]
        )

        data.append([
            str(idx),
            i.date.strftime("%d-%m-%Y"),
            i.semester,
            i.class_year,
            agenda_paragraph,
            ai_summary_paragraph,
            Paragraph(mentees_formatted, styles["Normal"]),
        ])

    # âœ… COLUMN WIDTHS (PRINT OPTIMIZED)
    table = RLTable(
        data,
        colWidths=[33, 56, 28, 25, 150, 150, 145],  # âœ… wider AI + Agenda
        repeatRows=1
    )

    # âœ… TABLE STYLING
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

        # âœ… Vertical alignment
        ("VALIGN", (0, 0), (-1, -1), "TOP"),

        # âœ… AUTO WORD WRAP (CRITICAL FIX)
        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),

        # âœ… Font control
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),

        # âœ… Padding (PREVENT OVERLAP)
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    story.append(table)
    story.append(Spacer(1, 40))
    # âœ… ATTENDANCE SUMMARY TABLE (PER MENTEE)

    # âœ… Get unique mentees from queryset
    attendance_map = {}

    total_interactions = qs.count()

    for interaction in qs:
        for u in interaction.mentees.all():
            if hasattr(u, "profile"):
                key = u.profile.moodle_id
                if key not in attendance_map:
                    attendance_map[key] = {
                        "name": u.profile.student_name,
                        "dept": u.profile.branch if u.profile.branch else "N/A",
                        "moodle": u.profile.moodle_id,
                        "present": 0,
                    }
                attendance_map[key]["present"] += 1

    # âœ… Build Attendance Table Data
    attendance_data = [["Sr. No.", "Name", "Department", "Moodle ID", "Present", "Attendance %"]]

    for idx, v in enumerate(attendance_map.values(), start=1):
        percent = round((v["present"] / total_interactions) * 100, 2) if total_interactions else 0
        attendance_data.append([
            str(idx),
            v["name"],
            v["dept"],
            v["moodle"],
            str(v["present"]),
            f"{percent}%",
        ])
    story.append(Spacer(1, 25))

    attendance_title = Paragraph("<b>Mentee Attendance Summary</b>", styles["Heading2"])
    story.append(attendance_title)
    story.append(Spacer(1, 10))

    attendance_table = RLTable(attendance_data, colWidths=[40, 130, 100, 100, 70, 90])

    attendance_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(attendance_table)

    # âœ… SIGNATURE SECTION
    sign_table = RLTable(
        [
            ["", ""],
            ["______________________________", "______________________________"],
            ["Mentor's Signature", "HOD / Coordinator Signature"]
        ],
        colWidths=[250, 250]
    )

    sign_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 20),
    ]))

    story.append(sign_table)

    # âœ… LOGO PATH
    logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")

    # âœ… FOOTER + PAGE NUMBER + LOGO
    def add_footer_and_page_number(canvas, doc):
        canvas.saveState()
        width, height = A4
        # âœ… WHITE BACKGROUND FOR LOGO
        canvas.setFillColor(colors.white)
        canvas.rect(30, height - 80, 90, 50, fill=1, stroke=0)

        # âœ… LOGO
        if os.path.exists(logo_path):
            canvas.drawImage(
                logo_path,
                35,
                height - 85,
                width=100,
                height=60,
                preserveAspectRatio=True,
                mask="auto"
            )

        # âœ… MAIN TITLE (FIRST PAGE)
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(colors.black)

        if doc.page == 1:
            canvas.drawString(130, height - 50, "Mentorâ€“Mentee Interaction Report")
        else:
            canvas.drawString(130, height - 50, "Mentorâ€“Mentee Interaction Report")

        # âœ… HEADER DETAILS
        canvas.setFont("Helvetica", 10)
        header_y = height - 70
        canvas.drawString(130, header_y, f"Department: {department}")
        canvas.drawString(350, header_y, f"Semester: {semester_filter}")
        canvas.drawString(130, header_y - 15, f"Generated On: {export_date}")

        # âœ… FOOTER WITH MENTOR NAME
        mentor_name = mentor.name if mentor.name else mentor.user.get_full_name() or mentor.user.username
        canvas.setFont("Helvetica", 9)
        canvas.drawString(36, 30, f"Generated by Mentor: {mentor_name}")

        # âœ… PAGE NUMBER
        canvas.drawRightString(width - 36, 30, f"Page {doc.page}")

        canvas.restoreState()

    # âœ… BUILD DOCUMENT
    doc.build(
        story,
        onFirstPage=add_footer_and_page_number,
        onLaterPages=add_footer_and_page_number
    )

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Mentor_Mentee_interactions.pdf")
#----------------Mentor-Mentee interaction page logic ends-----------------

#-----------------Student data visualization logic starts------------------
@login_required
def student_visualization(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)

    mentee_user_ids = (
        MentorMentee.objects
        .filter(mentor=mentor_obj)
        .values_list("mentee__user_id", flat=True)
    )

    internship_data = list(
        InternshipPBL.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )

    project_data = list(
        Project.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )

    sports_data = list(
        SportsCulturalEvent.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )

    course_data = list(
        CertificationCourse.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )

    hackathon_data = list(
        OtherEvent.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )

    paper_data = list(
        PaperPublication.objects
        .filter(user_id__in=mentee_user_ids)
        .values("user__profile__branch")
        .annotate(total=Count("id"))
    )
    # âœ… Collect all academic years from all models
    year_sets = [
        InternshipPBL.objects.values_list("academic_year", flat=True),
        Project.objects.values_list("academic_year", flat=True),
        SportsCulturalEvent.objects.values_list("academic_year", flat=True),
        CertificationCourse.objects.values_list("academic_year", flat=True),
        OtherEvent.objects.values_list("academic_year", flat=True),
        PaperPublication.objects.values_list("academic_year", flat=True),
    ]

    academic_years = sorted(
        set(y for y in chain(*year_sets) if y),
        reverse=True
    )

    return render(request, "mentor/student_visualization.html", {
        "internship_data": internship_data,
        "project_data": project_data,
        "sports_data": sports_data,
        "course_data": course_data,
        "hackathon_data": hackathon_data,
        "paper_data": paper_data,
        "academic_years": academic_years,
    })


@login_required
def get_chart_data(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)

    category = request.GET.get("category")
    academic_year = request.GET.get("year")

    model_map = {
        "internship": InternshipPBL,
        "project": Project,
        "sports": SportsCulturalEvent,
        "course": CertificationCourse,
        "other": OtherEvent,
        "paper": PaperPublication,
    }

    model = model_map.get(category)
    if not model:
        return JsonResponse({"labels": [], "data": []})

    # âœ… Only this mentorâ€™s mentees
    mentee_user_ids = (
        MentorMentee.objects
        .filter(mentor=mentor_obj)
        .values_list("mentee__user_id", flat=True)
    )

    qs = model.objects.filter(user_id__in=mentee_user_ids)

    if academic_year:
        qs = qs.filter(academic_year=academic_year)

    # âœ… GROUP BY STUDENT (not department)
    data = (
        qs.values(
            "user_id",
            "user__profile__student_name",
            "user__username",
            "user__profile__branch",
        )
        .annotate(count=Count("id"))
        .order_by("user__profile__student_name")
    )

    labels = [
        d["user__profile__student_name"]
        or d["user__username"]
        for d in data
    ]

    counts = [d["count"] for d in data]

    return JsonResponse({
        "labels": labels,
        "data": counts
    })



@login_required
def export_department_students_excel(request):
    # âœ… must be a mentor user
    mentor_obj = get_object_or_404(Mentor, user=request.user)

    category = request.GET.get("category")
    student_name = request.GET.get("student")
    academic_year = request.GET.get("year")

    if not category or not student_name:
        return HttpResponse("Missing category or student", status=400)

    model_map = {
        "internship": (InternshipPBL, "Internships_PBL"),
        "project": (Project, "Projects"),
        "sports": (SportsCulturalEvent, "Sports_Cultural"),
        "course": (CertificationCourse, "Courses_Certifications"),
        "other": (OtherEvent, "Other_Achievements"),
        "paper": (PaperPublication, "Paper_Publications"),
    }

    Model, category_label = model_map[category]

    mentee_users = User.objects.filter(
        id__in=MentorMentee.objects.filter(
            mentor=mentor_obj
        ).values_list("mentee__user_id", flat=True),
        profile__student_name=student_name
    )

    if not mentee_users.exists():
        return HttpResponse("Student not found", status=404)

    qs = Model.objects.filter(user__in=mentee_users)

    if academic_year:
        qs = qs.filter(academic_year=academic_year)

    qs = qs.order_by(
        "user__profile__student_name",
        "-uploaded_at" if hasattr(Model, "uploaded_at") else "id"
    )

    # âœ… Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "Moodle ID", "Student Name", "Department", "Semester", "Academic Year",
        "Record Title", "Type", "Details", "Uploaded At"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def safe(v):
        return "" if v is None else str(v)

    for obj in qs:
        profile = getattr(obj.user, "profile", None)

        moodle = safe(getattr(profile, "moodle_id", ""))
        name = safe(getattr(profile, "student_name", obj.user.get_full_name() or obj.user.username))
        branch = safe(getattr(profile, "branch", ""))
        semester = safe(getattr(obj, "semester", getattr(profile, "semester", "")))
        academic_year = safe(getattr(obj, "academic_year", ""))

        record_title = safe(getattr(obj, "title", None) or getattr(obj, "name_of_event", None) or "")
        record_type = safe(getattr(obj, "type", None) or getattr(obj, "project_type", None) or getattr(obj, "level", None) or "")

        details = safe(
            getattr(obj, "details", None)
            or getattr(obj, "company_name", None)
            or getattr(obj, "venue", None)
            or getattr(obj, "authors", None)
            or ""
        )

        uploaded_at = safe(getattr(obj, "uploaded_at", ""))

        ws.append([
            moodle, name, branch, semester, academic_year,
            record_title, record_type, details, uploaded_at
        ])

    # simple autosize
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    filename = f"{student_name}_{category_label}_MyMentees_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def get_chart_data_compare(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)
    category = request.GET.get("category")

    model_map = {
        "internship": InternshipPBL,
        "project": Project,
        "sports": SportsCulturalEvent,
        "course": CertificationCourse,
        "other": OtherEvent,
        "paper": PaperPublication,
    }

    model = model_map.get(category)
    if not model:
        return JsonResponse({"labels": [], "datasets": []})

    mentee_user_ids = (
        MentorMentee.objects
        .filter(mentor=mentor_obj)
        .values_list("mentee__user_id", flat=True)
    )

    qs = model.objects.filter(user_id__in=mentee_user_ids)

    years = sorted(qs.values_list("academic_year", flat=True).distinct())

    # âœ… STUDENTS instead of departments
    students = list(
        qs.values(
            "user_id",
            "user__profile__student_name",
            "user__username"
        ).distinct()
    )

    labels = years  # X-axis = years
    datasets = []

    for s in students:
        student_label = s["user__profile__student_name"] or s["user__username"]
        data = []

        for year in years:
            count = qs.filter(
                user_id=s["user_id"],
                academic_year=year
            ).count()
            data.append(count)

        datasets.append({
            "label": student_label,
            "data": data
        })

    return JsonResponse({
        "labels": labels,     # years on X-axis
        "datasets": datasets # one line per student
    })


@login_required
def get_heatmap_data(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)
    category = request.GET.get("category")

    model_map = {
        "internship": InternshipPBL,
        "project": Project,
        "sports": SportsCulturalEvent,
        "course": CertificationCourse,
        "other": OtherEvent,
        "paper": PaperPublication,
    }

    model = model_map.get(category)
    if not model:
        return JsonResponse({"years": [], "students": [], "matrix": []})

    mentee_user_ids = (
        MentorMentee.objects
        .filter(mentor=mentor_obj)
        .values_list("mentee__user_id", flat=True)
    )

    qs = model.objects.filter(user_id__in=mentee_user_ids)

    years = sorted(qs.values_list("academic_year", flat=True).distinct())

    students = list(
        qs.values(
            "user_id",
            "user__profile__student_name",
            "user__username"
        ).distinct()
    )

    student_labels = [
        s["user__profile__student_name"] or s["user__username"]
        for s in students
    ]

    matrix = []

    for year in years:
        row = []
        for s in students:
            row.append(
                qs.filter(
                    user_id=s["user_id"],
                    academic_year=year
                ).count()
            )
        matrix.append(row)

    return JsonResponse({
        "years": years,
        "students": student_labels,
        "matrix": matrix
    })


def signature_block(mentor_name):
    data = [
        ["", ""],
        ["__________________________", "__________________________"],
        [f"Mentor Signature\n{mentor_name}", "HOD Signature\nHead of Department"],
    ]

    table = RLTable(data, colWidths=[8 * cm, 8 * cm])

    table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONT", (0, 2), (-1, 2), "Helvetica-Bold"),
        ("TOPPADDING", (0, 1), (-1, 1), 20),
        ("TOPPADDING", (0, 2), (-1, 2), 6),
        ("TOPPADDING", (0, 3), (-1, -1), 10),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 10),
    ]))

    return table


def generate_performance_insights(user, mentor_user):
    # ================= SEMESTER RESULTS =================
    sem_results = SemesterResult.objects.filter(user=user).order_by("created_at")

    sgpas = [float(s.pointer) for s in sem_results if s.pointer]

    # ---------- SGPA TREND ----------
    trend = "No Data"
    if len(sgpas) >= 2:
        if sgpas[-1] > sgpas[0]:
            trend = "Improving"
        elif sgpas[-1] < sgpas[0]:
            trend = "Declining"
        else:
            trend = "Stable"

    # ---------- AVERAGE SGPA ----------
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else None

    # ---------- KT RISK ----------
    total_kts = sem_results.aggregate(total=Sum("no_of_kt"))["total"] or 0

    if total_kts == 0:
        kt_risk = "Low"
    elif total_kts <= 2:
        kt_risk = "Moderate"
    else:
        kt_risk = "High"

    # ================= ATTENDANCE =================
    # total sessions conducted by this mentor
    total_sessions = MentorMenteeInteraction.objects.filter(
        mentor=mentor_user
    ).count()

    # sessions where this student was present
    attended_sessions = MentorMenteeInteraction.objects.filter(
        mentor=mentor_user,
        mentees=user
    ).count()

    attendance_text = "N/A"
    attendance_pct = 0

    if total_sessions > 0:
        attendance_pct = round((attended_sessions / total_sessions) * 100, 1)

        if attendance_pct >= 80:
            attendance_text = f"Regular ({attendance_pct}%)"
        elif attendance_pct >= 60:
            attendance_text = f"Occasional ({attendance_pct}%)"
        else:
            attendance_text = f"Poor ({attendance_pct}%)"

    # ================= OVERALL SCORE =================
    score = 0

    if avg_sgpa:
        score += avg_sgpa * 10

    if total_kts == 0:
        score += 20

    if attendance_pct >= 80:
        score += 20
    elif attendance_pct >= 60:
        score += 10

    if score >= 85:
        overall = "Excellent"
    elif score >= 65:
        overall = "Good"
    else:
        overall = "Needs Attention"

    return {
        "trend": trend,
        "avg": avg_sgpa,
        "kt_risk": kt_risk,
        "attendance": attendance_text,
        "attendance_pct": attendance_pct,
        "overall": overall
    }


@login_required
def export_progress_pdf(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)
    mentor_name = mentor_obj.name or request.user.username

    mentees = MentorMentee.objects.filter(
        mentor=mentor_obj
    ).select_related("mentee__user", "mentee__user__profile")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=Mentor_Mentee_Progress.pdf"

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = []

    # =====================================================
    # HEADER / FOOTER
    # =====================================================
    logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")
    today = datetime.now().strftime("%d %b %Y")

    def draw_header_footer(canvas, doc):
        canvas.saveState()
        width, height = A4
        # ===== WHITE BG FOR LOGO =====
        canvas.setFillColor(colors.white)
        canvas.rect(30, height - 70, 90, 50, fill=1, stroke=0)

        # ===== LOGO =====
        if os.path.exists(logo_path):
            canvas.drawImage(
                logo_path,
                35,
                height - 80,
                width=100,
                height=60,
                preserveAspectRatio=True,
                mask="auto"  # enables PNG transparency
            )

        # ===== INSTITUTE NAME =====
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(colors.black)
        canvas.drawString(
                150,
                height - 50,
                "A. P. Shah Institute of Technology (APSIT)"
        )

        # ===== FOOTER =====
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawString(
            1.5 * cm,
            1 * cm,
            f"Mentor: {mentor_name}"
        )
        canvas.drawRightString(
            20 * cm,
            1 * cm,
            f"Mentee Progress Report | Generated on {today} | Page {doc.page}"
        )

        canvas.restoreState()

    # ================= MENTOR SUMMARY PAGE =================

    elements.append(Paragraph("Mentee Performance Summary", styles["Title"]))
    elements.append(Spacer(1, 12))

    summary_table = [[
        "Mentee Name",
        "Avg SGPA",
        "Internships",
        "Projects",
        "Events",
        "Courses",
        "Papers",
        "Other Events",
        "Total KTs"
    ]]

    for rel in mentees:
        user = rel.mentee.user
        profile = getattr(user, "profile", None)
        name = profile.student_name if profile else user.username

        avg_sgpa = SemesterResult.objects.filter(
            user=user
        ).aggregate(avg=Avg("pointer"))["avg"]

        total_kts = SemesterResult.objects.filter(
            user=user
        ).aggregate(total=Sum("no_of_kt"))["total"] or 0

        summary_table.append([
            name,
            round(avg_sgpa, 2) if avg_sgpa else "-",
            InternshipPBL.objects.filter(user=user).count(),
            Project.objects.filter(user=user).count(),
            SportsCulturalEvent.objects.filter(user=user).count(),
            CertificationCourse.objects.filter(user=user).count(),
            PaperPublication.objects.filter(user=user).count(),
            OtherEvent.objects.filter(user=user).count(),
            total_kts
        ])

    table = RLTable(summary_table, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # =====================================================
    # COLLECT ALL YEARS
    # =====================================================
    year_sets = [
        InternshipPBL.objects.values_list("academic_year", flat=True),
        Project.objects.values_list("academic_year", flat=True),
        SportsCulturalEvent.objects.values_list("academic_year", flat=True),
        CertificationCourse.objects.values_list("academic_year", flat=True),
        PaperPublication.objects.values_list("academic_year", flat=True),
        OtherEvent.objects.values_list("academic_year", flat=True),
        SemesterResult.objects.values_list("academic_year", flat=True),
    ]

    all_years = sorted(set(y for y in chain(*year_sets) if y))

    # =====================================================
    # PER STUDENT PAGES
    # =====================================================
    for rel in mentees:
        user = rel.mentee.user
        profile = getattr(user, "profile", None)

        student_name = profile.student_name if profile else user.username
        branch = profile.branch if profile else "N/A"

        elements.append(Paragraph(
            f"{student_name} ({branch})",
            styles["Heading2"]
        ))
        elements.append(Spacer(1, 8))

        # ================= TABLE 1: ACTIVITY SUMMARY =================
        summary_table_data = [[
            "Year", "Semesters", "Internships", "Projects",
            "Events", "Courses", "Papers", "Other Events"
        ]]

        for year in all_years:
            sem_qs = SemesterResult.objects.filter(
                user=user, academic_year=year
            )

            semesters = list(sem_qs.values_list("semester", flat=True))
            sem_display = ", ".join(semesters) if semesters else "-"

            summary_table_data.append([
                year,
                sem_display,
                InternshipPBL.objects.filter(user=user, academic_year=year).count(),
                Project.objects.filter(user=user, academic_year=year).count(),
                SportsCulturalEvent.objects.filter(user=user, academic_year=year).count(),
                CertificationCourse.objects.filter(user=user, academic_year=year).count(),
                PaperPublication.objects.filter(user=user, academic_year=year).count(),
                OtherEvent.objects.filter(user=user, academic_year=year).count(),
            ])

        summary_table = RLTable(summary_table_data, repeatRows=1)
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 14))

        # ================= TABLE 2: SEMESTER PERFORMANCE =================
        sem_results = SemesterResult.objects.filter(
            user=user
        ).order_by("academic_year", "semester")

        sem_table = [["Year", "Semester", "SGPA", "KTs"]]
        spark_data = []
        x = 0

        for sr in sem_results:
            sgpa = float(sr.pointer) if sr.pointer else None
            kts = sr.no_of_kt or 0

            sem_table.append([
                sr.academic_year or "-",
                sr.semester or "-",
                f"{sgpa:.2f}" if sgpa else "-",
                kts
            ])

            if sgpa:
                spark_data.append((x, sgpa))
                x += 1

        sems = RLTable(sem_table, repeatRows=1)

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]

        for i in range(1, len(sem_table)):
            try:
                if float(sem_table[i][2]) >= 8:
                    style_cmds.append(("TEXTCOLOR", (2, i), (2, i), colors.green))
            except:
                pass

            if sem_table[i][3] > 0:
                style_cmds.append(("TEXTCOLOR", (3, i), (3, i), colors.red))

        sems.setStyle(TableStyle(style_cmds))

        elements.append(Paragraph("Semester Performance", styles["Heading3"]))
        elements.append(sems)
        elements.append(Spacer(1, 8))

        # ================= SGPA SPARKLINE =================
        if spark_data:
            from reportlab.graphics.shapes import String

            drawing = Drawing(400, 140)
            lp = LinePlot()
            lp.x = 40
            lp.y = 30
            lp.width = 320
            lp.height = 80
            lp.data = [spark_data]

            lp.lines[0].strokeColor = colors.darkblue
            lp.lines[0].strokeWidth = 2

            # Axes
            lp.yValueAxis.valueMin = 0
            lp.yValueAxis.valueMax = 10
            lp.yValueAxis.visible = True
            lp.xValueAxis.visible = False

            drawing.add(lp)

            # ðŸ· SGPA LABELS ABOVE POINTS
            for i, (x_val, y_val) in enumerate(spark_data):
                # Convert data coords â†’ drawing coords
                x_pos = lp.x + (i / max(len(spark_data) - 1, 1)) * lp.width
                y_pos = lp.y + ((y_val - lp.yValueAxis.valueMin) /
                                (lp.yValueAxis.valueMax - lp.yValueAxis.valueMin)) * lp.height

                label = String(
                    x_pos,
                    y_pos + 6,  # slightly above the point
                    f"{round(y_val, 2)}",
                    fontSize=7,
                    fillColor=colors.black,
                    textAnchor="middle"
                )
                drawing.add(label)

            elements.append(Paragraph("SGPA Trend", styles["Italic"]))
            elements.append(Spacer(1, 4))
            elements.append(drawing)

        # ================= PERFORMANCE INSIGHTS =================
        insights = generate_performance_insights(user, request.user)

        insight_text = f"""
        <b>Performance Insights</b><br/>
        ðŸ“ˆ SGPA Trend: {insights['trend']}<br/>
        ðŸŽ¯ Academic Strength: {'Strong' if insights['avg'] and insights['avg'] >= 8 else 'Moderate'} (Avg SGPA: {insights['avg'] or 'N/A'})<br/>
        âš ï¸ KT Risk: {insights['kt_risk']}<br/>
        ðŸ§‘â€ðŸ« Mentoring Attendance: {insights['attendance']}<br/>
        ðŸ† Overall Performance: <b>{insights['overall']}</b><br/><br/>
        <b>Mentor Note:</b> ________________________________________________
        """

        elements.append(Paragraph(insight_text, styles["Normal"]))

        elements.append(Spacer(1, 20))

    # ================= FINAL SIGNATURE PAGE =================
    elements.append(PageBreak())

    mentor_name = request.user.get_full_name() or request.user.username

    elements.append(Spacer(1, 200))

    elements.append(Paragraph("Signatures", styles["Title"]))
    elements.append(Spacer(1, 40))

    sig_table_data = [
        ["__________________________", "__________________________"],
        [f"Mentor Signature\n{mentor_obj.name}", "HOD Signature\nHead of Department"],
        [f"Date: {datetime.now().strftime('%d %b %Y')}", ""]
    ]

    sig_table = RLTable(sig_table_data, [8 * cm, 8 * cm])
    sig_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONT", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, 0), 20),
        ("TOPPADDING", (0, 1), (-1, 1), 10),
        ("TOPPADDING", (0, 2), (-1, 2), 10),
    ]))

    elements.append(sig_table)

    doc.build(elements, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    return response


@login_required
def export_progress_excel(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)
    mentor_name = mentor_obj.name or request.user.username

    mentees = MentorMentee.objects.filter(
        mentor=mentor_obj
    ).select_related("mentee__user", "mentee__user__profile")

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="F4CCCC")

    # =====================================================
    # SHEET 1 â€” MENTOR SUMMARY
    # =====================================================
    ws = wb.active
    ws.title = "Mentor Summary"

    headers = [
        "Mentee Name",
        "Avg SGPA",
        "Internships",
        "Projects",
        "Events",
        "Courses",
        "Papers",
        "Other Events",
        "Total KTs"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for rel in mentees:
        user = rel.mentee.user
        profile = getattr(user, "profile", None)
        name = profile.student_name if profile else user.username

        avg_sgpa = SemesterResult.objects.filter(
            user=user
        ).aggregate(avg=Avg("pointer"))["avg"]

        total_kts = SemesterResult.objects.filter(
            user=user
        ).aggregate(total=Sum("no_of_kt"))["total"] or 0

        ws.append([
            name,
            round(avg_sgpa, 2) if avg_sgpa else "",
            InternshipPBL.objects.filter(user=user).count(),
            Project.objects.filter(user=user).count(),
            SportsCulturalEvent.objects.filter(user=user).count(),
            CertificationCourse.objects.filter(user=user).count(),
            PaperPublication.objects.filter(user=user).count(),
            OtherEvent.objects.filter(user=user).count(),
            total_kts
        ])

    # Auto width
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 35)

    # =====================================================
    # COLLECT YEARS
    # =====================================================
    year_sets = [
        InternshipPBL.objects.values_list("academic_year", flat=True),
        Project.objects.values_list("academic_year", flat=True),
        SportsCulturalEvent.objects.values_list("academic_year", flat=True),
        CertificationCourse.objects.values_list("academic_year", flat=True),
        PaperPublication.objects.values_list("academic_year", flat=True),
        OtherEvent.objects.values_list("academic_year", flat=True),
        SemesterResult.objects.values_list("academic_year", flat=True),
    ]

    all_years = sorted(set(y for y in chain(*year_sets) if y))

    # =====================================================
    # PER STUDENT SHEETS
    # =====================================================
    for rel in mentees:
        user = rel.mentee.user
        profile = getattr(user, "profile", None)

        student_name = profile.student_name if profile else user.username
        branch = profile.branch if profile else "N/A"

        sheet = wb.create_sheet(student_name[:31])

        # ---------- TITLE ----------
        sheet["A1"] = f"{student_name} ({branch})"
        sheet["A1"].font = Font(bold=True, size=14)

        row = 3

        # ================= ACTIVITY SUMMARY =================
        sheet[f"A{row}"] = "Yearly Activity Summary"
        sheet[f"A{row}"].font = Font(bold=True)
        row += 1

        headers = [
            "Year", "Semesters", "Internships",
            "Projects", "Events", "Courses",
            "Papers", "Other Events"
        ]

        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row += 1

        for year in all_years:
            sems = SemesterResult.objects.filter(
                user=user, academic_year=year
            ).values_list("semester", flat=True)

            sem_display = ", ".join(sems) if sems else ""

            sheet.append([
                year,
                sem_display,
                InternshipPBL.objects.filter(user=user, academic_year=year).count(),
                Project.objects.filter(user=user, academic_year=year).count(),
                SportsCulturalEvent.objects.filter(user=user, academic_year=year).count(),
                CertificationCourse.objects.filter(user=user, academic_year=year).count(),
                PaperPublication.objects.filter(user=user, academic_year=year).count(),
                OtherEvent.objects.filter(user=user, academic_year=year).count(),
            ])
            row += 1

        row += 2

        # ================= SEMESTER PERFORMANCE =================
        sheet[f"A{row}"] = "Semester Performance"
        sheet[f"A{row}"].font = Font(bold=True)
        row += 1

        headers = ["Year", "Semester", "SGPA", "KTs"]
        sheet.append(headers)

        for col in range(1, 5):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row += 1

        sem_results = SemesterResult.objects.filter(
            user=user
        ).order_by("academic_year", "semester")

        for sr in sem_results:
            sgpa = float(sr.pointer) if sr.pointer else ""
            kts = sr.no_of_kt or 0

            sheet.append([
                sr.academic_year or "",
                sr.semester or "",
                sgpa,
                kts
            ])

            # Highlight rules
            if sgpa and sgpa >= 8:
                sheet.cell(row=row, column=3).fill = green

            if kts > 0:
                sheet.cell(row=row, column=4).fill = red

            row += 1

        row += 2

        # ================= PERFORMANCE INSIGHTS =================
        insights = generate_performance_insights(user, request.user)

        sheet[f"A{row}"] = "Performance Insights"
        sheet[f"A{row}"].font = Font(bold=True)
        row += 1

        insight_text = (
            f"SGPA Trend: {insights['trend']}\n"
            f"Academic Strength: {'Strong' if insights['avg'] and insights['avg'] >= 8 else 'Moderate'} "
            f"(Avg SGPA: {insights['avg'] or 'N/A'})\n"
            f"KT Risk: {insights['kt_risk']}\n"
            f"Mentoring Attendance: {insights['attendance']}\n"
            f"Overall Performance: {insights['overall']}\n\n"
            f"Mentor Note: ________________________________"
        )

        sheet[f"A{row}"] = insight_text
        sheet.row_dimensions[row].height = 90

        # Auto width
        for col in sheet.columns:
            width = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    # =====================================================
    # EXPORT
    # =====================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Mentor_Mentee_Progress.xlsx"'
    )

    return response


@login_required
def get_department_trends(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)
    category = request.GET.get("category")

    model_map = {
        "internship": InternshipPBL,
        "project": Project,
        "sports": SportsCulturalEvent,
        "course": CertificationCourse,
        "other": OtherEvent,
        "paper": PaperPublication,
    }

    model = model_map.get(category)
    if not model:
        return JsonResponse({})

    mentee_user_ids = MentorMentee.objects.filter(
        mentor=mentor_obj
    ).values_list("mentee__user_id", flat=True)

    qs = model.objects.filter(user_id__in=mentee_user_ids)

    years = sorted(qs.values_list("academic_year", flat=True).distinct())

    if len(years) < 2:
        return JsonResponse({})

    current = years[-1]
    previous = years[-2]

    trends = {}
    percentages = {}

    students = qs.values(
        "user_id",
        "user__profile__student_name",
        "user__username"
    ).distinct()

    for s in students:
        student_label = s["user__profile__student_name"] or s["user__username"]

        curr_count = qs.filter(
            user_id=s["user_id"],
            academic_year=current
        ).count()

        prev_count = qs.filter(
            user_id=s["user_id"],
            academic_year=previous
        ).count()

        if curr_count > prev_count:
            trends[student_label] = "up"
        elif curr_count < prev_count:
            trends[student_label] = "down"
        else:
            trends[student_label] = "same"

        percentages[student_label] = round(
            ((curr_count - prev_count) / max(prev_count, 1)) * 100,
            1
        )

    return JsonResponse({
        "current": current,
        "previous": previous,
        "trends": trends,
        "percentages": percentages
    })


@login_required
@csrf_exempt
def export_filtered_progress_pdf(request):
    mentor_obj = get_object_or_404(Mentor, user=request.user)

    if request.method != "POST":
        return HttpResponse("Use POST to export PDF", status=405)

    try:
        raw = request.body.decode("utf-8")
        data = json.loads(raw)
    except Exception:
        return HttpResponse("Invalid JSON body", status=400)

    year = data.get("year")
    category = data.get("category")
    chart_b64 = data.get("chart")

    if not all([year, category, chart_b64]):
        return HttpResponse("Missing data", status=400)

    # ================= DECODE CHART =================
    try:
        if "," in chart_b64:
            chart_b64 = chart_b64.split(",")[1]
        chart_bytes = base64.b64decode(chart_b64)
    except Exception:
        return HttpResponse("Invalid chart image", status=400)

    model_map = {
        "internship": (InternshipPBL, "Internships / PBL"),
        "project": (Project, "Projects"),
        "sports": (SportsCulturalEvent, "Sports / Cultural"),
        "course": (CertificationCourse, "Courses / Certifications"),
        "other": (OtherEvent, "Other Achievements"),
        "paper": (PaperPublication, "Paper Publications"),
    }

    if category not in model_map:
        return HttpResponse("Invalid category", status=400)

    Model, category_label = model_map[category]

    mentees = MentorMentee.objects.filter(
        mentor=mentor_obj
    ).select_related("mentee__user", "mentee__user__profile")

    # ================= PDF SETUP =================
    response = HttpResponse(content_type="application/pdf")
    filename = f"{category_label}_{year}_Mentee_Report.pdf".replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # ================= HEADER / FOOTER =================
    logo_path = os.path.join(settings.MEDIA_ROOT, "logo.png")
    mentor_name = request.user.get_full_name() or request.user.username
    today = datetime.now().strftime("%d %b %Y")

    def draw_header_footer(canvas, doc):
        canvas.saveState()
        width, height =A4

        # Logo
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 1 * cm, 27 * cm, width=100, height=60, preserveAspectRatio=True, mask="auto")
            canvas.setFont("Helvetica-Bold", 10)
            # ===== INSTITUTE NAME =====
            canvas.setFont("Helvetica-Bold", 16)
            canvas.setFillColor(colors.black)
            canvas.drawString(
                150,
                height - 50,
                "A. P. Shah Institute of Technology (APSIT)"
            )

        # Footer
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            20 * cm,
            1 * cm,
            f"{mentor_obj.name} | Generated {today} | Page {doc.page}"
        )

        canvas.restoreState()

    # ================= COVER PAGE =================
    elements.append(Paragraph("Academic Progress Report", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"Category: {category_label} | Academic Year: {year}",
        styles["Italic"]
    ))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Performance Overview Chart", styles["Heading3"]))
    elements.append(Spacer(1, 6))
    chart_img = RLImage(BytesIO(chart_bytes))
    chart_img.drawWidth = 16 * cm
    chart_img.drawHeight = 9 * cm

    elements.append(chart_img)
    elements.append(PageBreak())

    # ================= STUDENT SECTIONS =================
    for rel in mentees:
        user = rel.mentee.user
        profile = getattr(user, "profile", None)

        name = profile.student_name if profile else user.username
        branch = profile.branch if profile else "N/A"

        qs = Model.objects.filter(
            user=user,
            academic_year=year
        )

        if not qs.exists():
            continue

        elements.append(Paragraph(
            f"{name} ({branch})",
            styles["Heading2"]
        ))
        elements.append(Spacer(1, 6))

        # ---------- COLUMN SET ----------
        if category in ["internship", "course"]:
            headers = ["Title", "Company / Authority", "Semester", "Start Date", "End Date", "Days"]
        elif category == "project":
            headers = ["Title", "Type", "Semester", "Guide Name", "Link"]
        elif category in ["sports", "other"]:
            headers = ["Event Name", "Semester", "Level", "Prize Won"]
        else:
            headers = ["Title", "Type", "Level", "Authors", "Semester"]

        table_data = [headers]

        for obj in qs:
            if category == "internship":
                table_data.append([
                    obj.title or "",
                    obj.company_name or "",
                    obj.semester or "-",
                    obj.start_date.strftime("%d-%b-%Y") if obj.start_date else "-",
                    obj.end_date.strftime("%d-%b-%Y") if obj.end_date else "-",
                    obj.no_of_days or "-"
                ])
            elif category == "course":
                table_data.append([
                    obj.title or "",
                    obj.certifying_authority or "",
                    obj.semester or "-",
                    obj.start_date.strftime("%d-%b-%Y") if obj.start_date else "-",
                    obj.end_date.strftime("%d-%b-%Y") if obj.end_date else "-",
                    obj.no_of_days or "-"
                ])
            elif category == "project":
                table_data.append([
                    obj.title or "",
                    obj.project_type or "",
                    obj.semester or "-",
                    obj.guide_name or "",
                    obj.link or ""
                ])
            elif category == "sports":
                table_data.append([
                    obj.name_of_event or "",
                    obj.semester or "-",
                    obj.level or "",
                    obj.prize_won or ""
                ])
            elif category == "other":
                table_data.append([
                    obj.name_of_event or "",
                    obj.semester or "-",
                    obj.level or "",
                    obj.prize_won or ""
                ])
            else:
                table_data.append([
                    obj.title or "",
                    obj.type or "",
                    obj.level or "",
                    obj.authors or "",
                    obj.semester or "-"
                ])

        table = RLTable(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ]))

        elements.append(table)
        # elements.append(PageBreak())

    # ================= SIGNATURE PAGE =================
    elements.append(Paragraph("Verified & Signed by", styles["Title"]))
    elements.append(Spacer(1, 40))

    sign_table = RLTable([
        ["Mentor Signature", "HOD Signature"],
        ["\n\n_________________________", "\n\n_________________________"],
        [mentor_obj.name, "Head of Department"]
    ], colWidths=[9 * cm, 9 * cm])

    sign_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TOPPADDING", (0, 1), (-1, 1), 30),
    ]))

    elements.append(sign_table)

    # ================= BUILD =================
    doc.build(elements, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    return response


@login_required
@csrf_exempt
def export_filtered_progress_excel(request):
    # ---------- SECURITY ----------
    if request.method != "POST":
        return HttpResponse("POST required", status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON", status=400)

    year = data.get("year")
    category = data.get("category")
    chart_b64 = data.get("chart")

    if not year or not category:
        return HttpResponse("Missing filters", status=400)

    mentor = get_object_or_404(Mentor, user=request.user)

    # ---------- CATEGORY MAP ----------
    model_map = {
        "internship": (InternshipPBL, ["company_name", "start_date", "end_date", "no_of_days", "type"]),
        "project": (Project, ["guide_name", "project_type", "link"]),
        "sports": (SportsCulturalEvent, ["level", "prize_won", "venue", "type"]),
        "course": (CertificationCourse, ["certifying_authority", "start_date", "end_date", "no_of_days", "domain", "level"]),
        "other": (OtherEvent, ["level", "prize_won", "details"]),
        "paper": (PaperPublication, ["type", "level", "conf_name", "authors"]),
    }

    if category not in model_map:
        return HttpResponse("Invalid category", status=400)

    Model, extra_fields = model_map[category]

    # ---------- MENTEES ----------
    mentee_user_ids = MentorMentee.objects.filter(
        mentor=mentor
    ).values_list("mentee__user_id", flat=True)

    qs = Model.objects.filter(
        user_id__in=mentee_user_ids,
        academic_year=year
    ).select_related("user", "user__profile")

    # ---------- EXCEL ----------
    wb = openpyxl.Workbook()

    # ================= SHEET 1: RAW DATA =================
    ws = wb.active
    ws.title = "Student Records"

    headers = [
        "Student Name",
        "Department",
        "Semester",
        "Academic Year",
        "Title / Event"
    ]

    headers += [h.replace("_", " ").title() for h in extra_fields]
    headers.append("Uploaded At")

    ws.append(headers)

    # ---------- HEADER STYLE ----------
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def safe(v):
        return "" if v is None else str(v)

    student_totals = {}
    student_moodle = {}

    for obj in qs:
        profile = getattr(obj.user, "profile", None)
        student = profile.student_name if profile else obj.user.username
        moodle_id = getattr(profile, "moodle_id", "")
        student_moodle[student] = moodle_id

        student_totals[student] = student_totals.get(student, 0) + 1

        row = [
            student,
            safe(getattr(profile, "branch", "")),
            safe(getattr(obj, "semester", "")),
            safe(getattr(obj, "academic_year", "")),
            safe(
                getattr(obj, "title", None)
                or getattr(obj, "name_of_event", None)
                or ""
            ),
        ]

        for f in extra_fields:
            row.append(safe(getattr(obj, f, "")))

        row.append(safe(getattr(obj, "uploaded_at", "")))
        ws.append(row)

    # ================= COLOR CODE ROWS =================
    green = PatternFill("solid", fgColor="C6EFCE")   # 10+
    yellow = PatternFill("solid", fgColor="FFF2CC") # 5â€“10
    red = PatternFill("solid", fgColor="F4CCCC")    # 0â€“5
    #
    # for r in range(2, ws.max_row + 1):
    #     student_name = ws.cell(row=r, column=1).value
    #     total = student_totals.get(student_name, 0)
    #
    #     fill = green if total >= 10 else yellow if total >= 5 else red
    #
    #     for c in range(1, ws.max_column + 1):
    #         ws.cell(row=r, column=c).fill = fill

    # ================= AUTO WIDTH =================
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 3, 40)

    # ================= SHEET 2: SUMMARY =================
    pivot = wb.create_sheet("Summary")

    pivot_headers = ["Moodle ID", "Student Name", "Total Records", "Performance"]
    pivot.append(pivot_headers)

    for col in range(1, 5):
        cell = pivot.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for student, total in sorted(student_totals.items()):
        label = "Excellent" if total >= 10 else "Moderate" if total >= 5 else "Needs Attention"
        pivot.append([student_moodle.get(student, ""),student, total, label])

    # ---------- COLOR SUMMARY ----------
    for r in range(2, pivot.max_row + 1):
        total = pivot.cell(row=r, column=3).value
        try:
            total = int(total)
        except:
            total = 0
        fill = green if total >= 10 else yellow if total >= 5 else red

        for c in range(1, 5):
            pivot.cell(row=r, column=c).fill = fill

    # ================= SHEET 3: CHART DATA + LIVE CHART =================
    chart_sheet = wb.create_sheet("Performance Chart")

    # ---------- TABLE HEADER ----------
    chart_sheet.append(["Moodle ID", "Student", "Total Records", "Category", "Year"])

    # ---------- TABLE DATA ----------
    for student, total in sorted(student_totals.items()):
        chart_sheet.append([
            student_moodle.get(student, ""),  # Moodle ID
            student,
            total,
            category.title(),
            year
        ])

    # ---------- AUTO WIDTH ----------
    for col in chart_sheet.columns:
        width = max(len(str(c.value or "")) for c in col)
        chart_sheet.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    # ---------- TITLE ----------
    chart_sheet["G1"] = f"{category.title()} - {year} Performance Chart"
    chart_sheet["G1"].font = Font(bold=True)

    # ================= LIVE BAR CHART =================
    bar_chart = BarChart()
    bar_chart.title = "Student Performance"
    bar_chart.y_axis.title = "Total Records"
    bar_chart.x_axis.title = "Students"

    # Data range
    data = Reference(
        chart_sheet,
        min_col=3,  # Total Records
        min_row=1,
        max_row=chart_sheet.max_row
    )

    # Category labels
    cats = Reference(
        chart_sheet,
        min_col=2,  # Student Names
        min_row=2,
        max_row=chart_sheet.max_row
    )

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)

    # Show values on bars
    bar_chart.dataLabels = DataLabelList()
    bar_chart.dataLabels.showVal = True

    bar_chart.width = 22
    bar_chart.height = 12

    # ---------- PLACE CHART BESIDE TABLE ----------
    chart_sheet.add_chart(bar_chart, "G3")

    # ================= SHEET 5: TREND LINE CHART =================
    trend_sheet = wb.create_sheet("Trend Chart")

    trend_sheet["A1"] = "Student Trend Over Time"
    trend_sheet["A1"].font = Font(bold=True)

    # Build year-based trend table
    trend_data = {}
    for student in student_totals:
        trend_data.setdefault(student, {})
        trend_data[student][year] = student_totals[student]

    headers = ["Student", year]
    trend_sheet.append(headers)

    for student, data_map in trend_data.items():
        trend_sheet.append([student, data_map.get(year, 0)])

    line_chart = LineChart()
    line_chart.title = "Performance Trend"
    line_chart.y_axis.title = "Records"
    line_chart.x_axis.title = "Student"

    data = Reference(
        trend_sheet,
        min_col=2,
        min_row=1,
        max_row=trend_sheet.max_row
    )

    cats = Reference(
        trend_sheet,
        min_col=1,
        min_row=2,
        max_row=trend_sheet.max_row
    )

    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)

    line_chart.width = 22
    line_chart.height = 12

    trend_sheet.add_chart(line_chart, "D3")

    #===============SHEET 6: CATEGORY COMPARISON CHART================
    compare_sheet = wb.create_sheet("Category Compare")

    compare_sheet.append(["Category", "Total Records"])

    compare_models = {
        "Internship": InternshipPBL,
        "Project": Project,
        "Sports": SportsCulturalEvent,
        "Course": CertificationCourse,
        "Other": OtherEvent,
        "Paper": PaperPublication,
    }

    for cat_name, CatModel  in compare_models.items():
        total = CatModel.objects.filter(
            user_id__in=mentee_user_ids,
            academic_year=year
        ).count()
        compare_sheet.append([cat_name, total])

    cat_chart = BarChart()
    cat_chart.title = "Category Comparison"
    cat_chart.y_axis.title = "Total Records"
    cat_chart.x_axis.title = "Category"

    data = Reference(compare_sheet, min_col=2, min_row=1, max_row=compare_sheet.max_row)
    cats = Reference(compare_sheet, min_col=1, min_row=2, max_row=compare_sheet.max_row)

    cat_chart.add_data(data, titles_from_data=True)
    cat_chart.set_categories(cats)

    cat_chart.dataLabels = DataLabelList()
    cat_chart.dataLabels.showVal = True

    cat_chart.width = 18
    cat_chart.height = 10

    compare_sheet.add_chart(cat_chart, "D3")


    # ================= EXPORT =================
    filename = f"{category}_{year}_Filtered_Progress_Report.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

#-----------------Student data visualization logic ends------------------


#-----------------Agenda page logic strats-------------------
@mentor_or_staff_required
def weekly_agenda_page(request):
    # ---------- Filters (GET) ----------
    week = request.GET.get("week", "").strip()
    year = request.GET.get("year", "").strip()
    sem = request.GET.get("sem", "").strip()

    qs = WeeklyAgenda.objects.all()
    if week:
        qs = qs.filter(week=week)
    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(sem=sem)

    # ---------- Edit mode (GET ?edit=ID) ----------
    edit_id = request.GET.get("edit")
    entry = None
    if edit_id:
        entry = get_object_or_404(WeeklyAgenda, pk=edit_id)

    can_edit = request.user.is_staff or (hasattr(request.user, "mentor") and request.user.is_mentor)

    # helper: redirect back preserving filters
    def redirect_with_filters():
        params = {}
        if week: params["week"] = week
        if year: params["year"] = year
        if sem: params["sem"] = sem
        url = reverse("weekly-agenda")
        if params:
            url = f"{url}?{urlencode(params)}"
        return redirect(url)

    # ---------- POST actions ----------
    if request.method == "POST":
        action = request.POST.get("action")

        # âœ… PASTE DELETE ALL HERE (right after action)
        if action == "delete_all":
            if not request.user.is_staff:
                messages.error(request, "Only staff can delete all entries.")
                return redirect_with_filters()

            WeeklyAgenda.objects.all().delete()
            messages.success(request, "All entries deleted.")
            return redirect_with_filters()

        # DELETE
        if action == "delete":
            delete_id = request.POST.get("delete_id")
            obj = get_object_or_404(WeeklyAgenda, pk=delete_id)
            obj.delete()
            messages.success(request, "Entry deleted.")
            return redirect_with_filters()

        # CREATE / UPDATE
        entry_id = request.POST.get("entry_id")
        obj_instance = None
        if entry_id:
            obj_instance = get_object_or_404(WeeklyAgenda, pk=entry_id)

        form = WeeklyAgendaForm(request.POST, request.FILES, instance=obj_instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.pk:
                obj.created_by = request.user
            obj.save()
            messages.success(request, "Entry updated." if entry_id else "Entry created.")
            return redirect_with_filters()

    else:
        form = WeeklyAgendaForm(instance=entry)

    context = {
        "form": form,
        "entries": qs,
        "filters": {"week": week, "year": year, "sem": sem},
        "can_edit": can_edit,
        "editing": entry is not None,
        "editing_entry": entry,
        "WEEK_CHOICES": WEEK_CHOICES,
        "YEAR_CHOICES": YEAR_CHOICES,
        "SEM_CHOICES": SEM_CHOICES,
    }
    return render(request, "mentor/weekly_agenda.html", context)
#----------------------Agenda page logic ends---------------------