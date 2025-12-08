from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date
from mentee.models import ActivityLog
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import io
from django.utils import timezone
from ..signals import get_diff
from ..utils import to_ist
from django.db.models import F
from django.core.paginator import Paginator


@staff_member_required
def activity_logs_view(request):
    # Render main page — frontend will call /api/activity-logs/ for data
    return render(request, "admin/activity-logs.html", {})

@staff_member_required
def activity_logs_api(request):
    # Live filter endpoint (returns JSON)
    qs = ActivityLog.objects.all().order_by('-timestamp')

    user_q = request.GET.get("user")
    action_q = request.GET.get("action")
    module_q = request.GET.get("module")
    date_q = request.GET.get("date")
    search_q = request.GET.get("q")
    only_changed = request.GET.get('only_changed')

    if user_q:
        qs = qs.filter(user__username__icontains=user_q)
    if action_q:
        qs = qs.filter(action__icontains=action_q)
    if module_q:
        qs = qs.filter(module__icontains=module_q)
    if date_q:
        try:
            d = parse_date(date_q)
            if d:
                qs = qs.filter(timestamp__date=d)
        except:
            pass
    if search_q:
        qs = qs.filter(details__icontains=search_q)

    # --- NEW: server-side changed-only filter ---
    if only_changed == "1":
        qs = qs.filter(old_data__isnull=False)
        qs = qs.filter(new_data__isnull=False)
        qs = qs.exclude(old_data=F("new_data"))

    # pagination params (simple)
    page = int(request.GET.get("page", 1))
    per = int(request.GET.get("per", 50))
    paginator = Paginator(qs, per)
    page_obj = paginator.get_page(page)
    start = (page - 1) * per
    end = start + per

    items = []
    for lo in qs.order_by("-timestamp")[start:end]:
        items.append({
            "id": lo.id,
            "user": lo.user.username if lo.user else None,
            "action": lo.action,
            "module": lo.module,
            "details": lo.details,
            "changes": get_diff(lo.old_data, lo.new_data),
            "old_data": lo.old_data,
            "new_data": lo.new_data,
            "ip": lo.ip_address,
            "browser": lo.browser,
            "path": lo.request_path,
            "timestamp": lo.timestamp.isoformat(),
        })

    return JsonResponse({
        "results": items,
        "count": qs.count(),
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "total": paginator.count
    })

@staff_member_required
def export_logs_csv(request):
    qs = ActivityLog.objects.all().order_by("-timestamp")
    # apply same filters as api if provided
    user_q = request.GET.get("user")
    if user_q:
        qs = qs.filter(user__username__icontains=user_q)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="activity_logs.csv"'
    writer = csv.writer(response)
    writer.writerow(["id","user","action","module","details","ip","browser","path","timestamp","old_data","new_data"])
    for lo in qs:
        writer.writerow([
            lo.id,
            lo.user.username if lo.user else "",
            lo.action,
            lo.module,
            lo.details,
            lo.ip_address,
            lo.browser,
            lo.request_path,
            to_ist(lo.timestamp),
            (lo.old_data and str(lo.old_data)) or "",
            (lo.new_data and str(lo.new_data)) or "",
        ])
    return response

@staff_member_required
def export_logs_excel(request):
    qs = ActivityLog.objects.all().order_by("-timestamp")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    headers = ["id","user","action","module","details","ip","browser","path","timestamp","old_data","new_data"]
    ws.append(headers)
    for lo in qs:
        ws.append([
            lo.id,
            lo.user.username if lo.user else "",
            lo.action,
            lo.module,
            lo.details,
            lo.ip_address,
            lo.browser,
            lo.request_path,
            to_ist(lo.timestamp),
            (lo.old_data and str(lo.old_data)) or "",
            (lo.new_data and str(lo.new_data)) or "",
        ])
    # auto width
    for i, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > max_len:
                    max_len = len(v)
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 80)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="activity_logs.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def export_logs_pdf(request):
    logs = ActivityLog.objects.all().order_by('-timestamp')
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(30, height - 40, "Website Activity Logs")

    y = height - 80
    data = [["User", "Action", "Module", "IP", "Time"]]

    for log in logs:
        data.append([
            log.user.username if log.user else "Anonymous",
            log.action,
            log.module,
            log.ip_address or "-",
            to_ist(log.timestamp),
        ])

    table = Table(data, colWidths=[120,150,110,100,120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.3, colors.black),
        ("FONT", (0,0), (-1,-1), "Helvetica", 8),
    ]))
    table.wrapOn(p, width, height)
    table.drawOn(p, 30, y - 15 * len(data))

    # Footer
    p.setFont("Helvetica", 9)
    p.drawCentredString(width/2, 20, f"Generated on {timezone.now().strftime('%d-%m-%Y %H:%M')}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")
